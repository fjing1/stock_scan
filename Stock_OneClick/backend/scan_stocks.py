# scan_stocks.py
import pandas as pd
import numpy as np
import yfinance as yf
from pathlib import Path
from datetime import datetime
import subprocess
import traceback
import shutil
import re
import sys
import os
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from xunlong import XunLongIndicator

# ================= 基本配置 =================

# 以项目根目录为基准（脚本位于 backend/）
BASE_DIR = SCRIPT_DIR.parent
INPUT_FILE = BASE_DIR / "stock_input_template.xlsx"
HISTORY_DIR = BASE_DIR / "history"
EXPORT_DIR = BASE_DIR / "exports"
COMPLETED_20D_DIR = HISTORY_DIR / "completed_14d"

V1_LOOKBACK_DAYS = 5
V2_LOOKBACK_DAYS = 5
GANN_LOOKBACK_DAYS = 10
V2_MAX_RANK120 = 0.4  # V2 低位过滤：rank120 <= 0.4
TRACK_MAX_DAYS = 14
LIFECYCLE_START_DATE = pd.Timestamp("2026-05-22").date()
A_POOL_SYMBOLS = [
    "AAPL", "MSFT", "AMZN", "META", "NVDA", "AMD", "AVGO", "MU", "MRVL",
    "ANET", "VRT", "LITE", "COHR", "FN", "VICR", "PANW", "CRWD", "SNOW",
    "PLTR", "TSLA", "JPM", "FCX",
]

CUSTOM_WATCHLISTS_CN = [
    ("01 市场环境", ["SP:SPX", "NASDAQ:NDX", "CBOE:VIX", "TVC:DXY", "NASDAQ:TLT", "AMEX:GLD", "AMEX:SLV", "NYMEX:CL1!", "COMEX:HG1!"]),
    ("02 AI超级核心", ["NASDAQ:NVDA", "NASDAQ:MSFT", "NASDAQ:AMZN", "NASDAQ:META", "NASDAQ:GOOGL", "NASDAQ:TSLA"]),
    ("03 GPU / AI芯片", ["NASDAQ:NVDA", "NASDAQ:AMD", "NASDAQ:AVGO", "NASDAQ:MU", "NASDAQ:INTC"]),
    ("04 半导体设备 / EDA", ["NASDAQ:ASML", "NASDAQ:AMAT", "NASDAQ:KLAC", "NASDAQ:LRCX", "NASDAQ:SNPS", "NASDAQ:CDNS"]),
    ("05 AI网络 / 光模块", ["NYSE:ANET", "NASDAQ:LITE", "NYSE:COHR", "NASDAQ:MRVL", "NASDAQ:CRDO"]),
    ("06 服务器 / 数据中心硬件", ["NYSE:VRT", "NASDAQ:SMCI", "NYSE:DELL", "NYSE:HPE"]),
    ("07 数据中心电力", ["NYSE:VST", "NYSE:NRG", "NASDAQ:TLN", "NYSE:AES", "NYSE:GEV"]),
    ("08 核能", ["NASDAQ:CEG", "NYSE:SMR", "NYSE:OKLO", "NYSE:LEU", "NYSE:CCJ", "NYSE:BWXT", "AMEX:UEC", "AMEX:UUUU", "NASDAQ:NNE"]),
    ("09 储能", ["NASDAQ:FLNC", "NASDAQ:EOSE", "NYSE:BE", "NASDAQ:MVST", "NASDAQ:PLUG"]),
    ("10 太阳能", ["NASDAQ:FSLR", "NASDAQ:NXT", "NASDAQ:ENPH", "NASDAQ:RUN"]),
    ("11 AI软件 / SaaS", ["NASDAQ:PLTR", "NASDAQ:ADBE", "NYSE:CRM", "NYSE:IBM", "NASDAQ:SNOW", "NASDAQ:DDOG", "NASDAQ:MDB", "NASDAQ:TEAM", "NYSE:NOW"]),
    ("12 网络安全", ["NASDAQ:CRWD", "NASDAQ:FTNT", "NASDAQ:ZS", "NASDAQ:PANW", "NASDAQ:OKTA"]),
    ("13 材料 / 电网", ["NYSE:FCX", "NASDAQ:LIN", "NYSE:APD", "NYSE:NUE", "NYSE:RIO", "NYSE:BHP"]),
    ("14 能源", ["NYSE:CVX", "NYSE:COP", "NYSE:EOG", "NYSE:HAL", "NYSE:SLB"]),
    ("15 公用事业", ["NYSE:NEE", "NYSE:SO", "NYSE:DUK", "NYSE:D"]),
    ("16 数据中心REIT", ["NASDAQ:EQIX", "NYSE:AMT", "NYSE:CCI", "NYSE:PLD"]),
    ("17 高弹性交易池", ["NASDAQ:APLD", "NASDAQ:CRWV", "NASDAQ:SOLS"]),
]

TV_TO_YF_SYMBOL_MAP = {
    "SPX": "^GSPC",
    "NDX": "^NDX",
    "VIX": "^VIX",
    "DXY": "DX-Y.NYB",
    "TLT": "TLT",
    "GLD": "GLD",
    "SLV": "SLV",
    "CL1!": "CL=F",
    "HG1!": "HG=F",
}

MARKET_CONTEXT_SYMBOLS = [
    ("SPX", "S&P 500", "核心指数"),
    ("QQQ", "Nasdaq 100 ETF", "核心指数"),
    ("IWM", "Russell 2000 ETF", "核心指数"),
    ("RSP", "S&P 500 等权ETF", "市场宽度"),
    ("VIX", "波动率指数", "风险确认"),
    ("HYG", "高收益债ETF", "信用风险"),
    ("SMH", "半导体ETF", "进攻温度计"),
    ("XLU", "公用事业ETF", "防守温度计"),
]

MARKET_CONTEXT_HELP_TEXT = """市场方向判断方法（写死在程序内，不依赖主交易清单）

固定观察标的：
核心指数：SPX / QQQ / IWM
市场宽度：RSP
风险确认：VIX / HYG
进攻温度计：SMH
防守温度计：XLU

主方向原则：
1. 主状态用日线决定，不用4H直接改主方向。
2. 4H只作为短线风险提示。
3. RSP用于判断市场宽度，HYG用于判断信用风险。
4. SMH代表进攻资金，XLU代表防守资金。

风险加分规则：
- SPX、QQQ、IWM 日线正式卖出：增加风险。
- SPX、QQQ、IWM 跌破MA20：增加风险。
- SPX或QQQ单日大跌：增加风险。
- RSP弱于SPX且跌破MA20：提示内部宽度变差。
- HYG跌破MA20或短期走弱：提示信用风险升高。
- VIX单日大涨或站上MA20：提示恐慌上升。
- SMH转弱而XLU强于SMH：提示资金偏防守。

输出档位：
强势/中性看涨：正常观察买点。
谨慎看涨：买点可看，但仓位和追涨要收敛。
风险升高：买点降权，优先高分、防守、现金流。
看跌/避险：暂停追高，高beta和进攻板块只观察。

当前版本只做Excel顶部提示，不直接修改个股观海买点分。
"""


# ================= 通用处理 =================

def normalize_yf_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    yfinance 单票时可能返回 MultiIndex 列（Price/Ticker）。
    这里把第二层 ticker 去掉，只保留 Price 这一层，便于后续处理。
    """
    if isinstance(df.columns, pd.MultiIndex):
        df = df.copy()
        df.columns = df.columns.get_level_values(0)
    return df


# ================= 工具函数 =================

def load_input_and_meta(path: Path):
    """读取 Sheet1_Input & Sheet2_Classified"""
    xls = pd.ExcelFile(path)
    try:
        df_input = pd.read_excel(xls, sheet_name="Sheet1_Input")
    except ValueError:
        df_input = pd.DataFrame(columns=["symbol"])

    try:
        df_meta = pd.read_excel(xls, sheet_name="Sheet2_Classified")
    except ValueError:
        df_meta = pd.DataFrame(columns=[
            "symbol", "name", "exchange", "sector",
            "industry", "market_cap", "group", "note", "enable"
        ])

    df_input["symbol"] = (
        df_input["symbol"].astype(str).str.strip().str.upper()
    )
    df_input = df_input[df_input["symbol"] != ""]

    if not df_meta.empty:
        df_meta["symbol"] = df_meta["symbol"].astype(str).str.strip().str.upper()

    return df_input, df_meta


def exchange_to_tv_prefix(exchange: str) -> str:
    """
    把常见 yfinance 交易所代码映射为 TradingView 前缀。
    未识别时返回空字符串，后续回退为纯 symbol。
    """
    ex = str(exchange or "").strip().upper()
    if ex in {"NMS", "NAS", "NASDAQ", "NCM", "NGM"}:
        return "NASDAQ"
    if ex in {"NYQ", "NYS", "NYSE"}:
        return "NYSE"
    if ex in {"ASE", "AMEX", "PCX", "ARCX"}:
        return "AMEX"
    return ""


def build_tv_symbol(symbol: str, exchange: str) -> str:
    prefix = exchange_to_tv_prefix(exchange)
    if prefix:
        return f"{prefix}:{symbol}"
    return symbol


def to_yfinance_symbol(symbol: str) -> str:
    """
    将输入池/TradingView 风格代码转换为 yfinance 可识别代码。
    """
    sym = str(symbol or "").strip().upper()
    if ":" in sym:
        sym = sym.split(":", 1)[1]
    return TV_TO_YF_SYMBOL_MAP.get(sym, sym)


def _extract_anchor_signals(history_dir: Path, current_df: pd.DataFrame, signal_side: str = "BUY") -> pd.DataFrame:
    """
    从历史 scan_result 文件 + 当天 df_all 中提取锚点信号。
    """
    frames = []
    signal_side = str(signal_side or "").upper()
    pat = re.compile(r"^scan_result_\d{8}_\d{6}\.xlsx$")
    files = sorted([p for p in history_dir.glob("scan_result_*.xlsx") if pat.match(p.name)])
    for p in files:
        try:
            tmp = _read_signal_rows_from_result(p, signal_side)
        except Exception:
            continue
        if {"symbol", "signal_date", "close"}.issubset(set(tmp.columns)):
            cols = ["symbol", "signal_date", "close"]
            if "signal_type" in tmp.columns:
                cols.append("signal_type")
            if "buy_score" in tmp.columns:
                cols.append("buy_score")
            frames.append(tmp[cols].copy())

    if not current_df.empty and {"symbol", "signal_date", "close"}.issubset(set(current_df.columns)):
        cur = current_df.copy()
        if "signal_side" in cur.columns:
            cur = cur[cur["signal_side"].astype(str).str.upper() == signal_side]
        cols = ["symbol", "signal_date", "close"]
        if "signal_type" in cur.columns:
            cols.append("signal_type")
        if "buy_score" in cur.columns:
            cols.append("buy_score")
        frames.append(cur[cols].copy())

    if not frames:
        return pd.DataFrame(columns=["symbol", "signal_date", "d0_close", "d0_rule"])

    all_sig = pd.concat(frames, ignore_index=True)
    all_sig["symbol"] = all_sig["symbol"].astype(str).str.strip().str.upper()
    all_sig["signal_date"] = pd.to_datetime(all_sig["signal_date"], errors="coerce").dt.date
    all_sig["close"] = pd.to_numeric(all_sig["close"], errors="coerce")
    all_sig = all_sig.dropna(subset=["symbol", "signal_date", "close"])
    all_sig = all_sig[all_sig["signal_date"] >= LIFECYCLE_START_DATE]
    all_sig = all_sig[all_sig["symbol"] != ""]
    if "signal_type" not in all_sig.columns:
        all_sig["signal_type"] = ""
    all_sig["signal_type"] = all_sig["signal_type"].fillna("").astype(str)
    if "buy_score" not in all_sig.columns:
        all_sig["buy_score"] = np.nan
    all_sig["buy_score"] = pd.to_numeric(all_sig["buy_score"], errors="coerce")
    if signal_side == "SELL":
        all_sig = all_sig[all_sig["signal_type"] == "正式卖出"]
    if all_sig.empty:
        return pd.DataFrame(columns=["symbol", "signal_date", "d0_close", "d0_rule"])

    grouped = (
        all_sig.sort_values(["signal_date", "symbol", "signal_type"])
        .groupby(["symbol", "signal_date"], as_index=False)
        .agg(
            d0_close=("close", "first"),
            d0_rule=("signal_type", lambda s: " | ".join([x for x in pd.unique(s) if x])),
            buy_score=("buy_score", "max"),
        )
    )
    return grouped.reset_index(drop=True)


def _read_signal_rows_from_result(path: Path, signal_side: str) -> pd.DataFrame:
    """
    兼容两种历史结果格式：
    1. 老版 Summary：单一表头
    2. 新版 Summary：买入区 / 卖出区分段
    """
    signal_side = str(signal_side or "").upper()
    try:
        raw_sig = pd.read_excel(path, sheet_name="RawSignals")
    except Exception:
        raw_sig = pd.DataFrame()

    if not raw_sig.empty and {"symbol", "signal_date", "close"}.issubset(set(raw_sig.columns)):
        if "signal_side" in raw_sig.columns:
            raw_sig = raw_sig[raw_sig["signal_side"].astype(str).str.upper() == signal_side]
        if "buy_score" not in raw_sig.columns:
            raw_sig["buy_score"] = raw_sig.apply(score_buy_signal_row, axis=1)
        return raw_sig

    try:
        tmp = pd.read_excel(path, sheet_name=0)
    except Exception:
        return pd.DataFrame()

    if {"symbol", "signal_date", "close"}.issubset(set(tmp.columns)):
        if "signal_side" in tmp.columns:
            tmp = tmp[tmp["signal_side"].astype(str).str.upper() == signal_side]
        return tmp

    raw = pd.read_excel(path, sheet_name=0, header=None)
    if raw.empty:
        return pd.DataFrame()

    sections = []
    for i in range(len(raw) - 1):
        row0 = str(raw.iloc[i, 0]) if pd.notna(raw.iloc[i, 0]) else ""
        next_row = raw.iloc[i + 1].tolist()
        if "run_date" in [str(x) for x in next_row]:
            title = row0.strip()
            if signal_side == "BUY" and "买入" not in title:
                continue
            if signal_side == "SELL" and "卖出" not in title:
                continue
            header = [str(x) if pd.notna(x) else "" for x in next_row]
            data_rows = []
            j = i + 2
            while j < len(raw):
                first_val = raw.iloc[j, 0]
                if pd.isna(first_val):
                    break
                first_txt = str(first_val).strip()
                if first_txt in {"排名", "No signals"} or "板块Top5统计" in first_txt or "买入跟踪" in first_txt or "卖出跟踪" in first_txt:
                    break
                vals = raw.iloc[j, :len(header)].tolist()
                data_rows.append(vals)
                j += 1
            if data_rows:
                sec = pd.DataFrame(data_rows, columns=header)
                sections.append(sec)

    if not sections:
        return pd.DataFrame()
    out = pd.concat(sections, ignore_index=True)
    return out


def _write_raw_signals_sheet(writer, df_all: pd.DataFrame):
    raw_cols = [
        "run_date", "run_time",
        "symbol", "name", "板块",
        "signal_date", "signal_type", "signal_side", "model",
        "close", "volume", "vol_ma20",
        "L2_trend", "L2_pump", "RSI",
        "rank120", "H4_RSI", "H4_FJ", "H4_0_birth", "H4_1_birth",
        "Gann_1_date", "Gann_1_price", "buy_score", "sell_score", "extra_info",
    ]
    raw_df = df_all.copy() if df_all is not None else pd.DataFrame()
    for c in raw_cols:
        if c not in raw_df.columns:
            raw_df[c] = np.nan
    raw_df = raw_df[raw_cols]
    raw_df.to_excel(writer, sheet_name="RawSignals", index=False)


def export_signal_dashboard(df_all: pd.DataFrame, run_dt: datetime, export_dir: Path, history_dir: Path):
    """导出一个轻量 HTML dashboard，集中显示当前 BUY/SELL 信号。"""
    export_dir.mkdir(parents=True, exist_ok=True)
    history_dir.mkdir(parents=True, exist_ok=True)
    cur_path = export_dir / "xl_signal_dashboard_latest.html"
    hist_path = history_dir / f"xl_signal_dashboard_{run_dt.strftime('%Y%m%d_%H%M%S')}.html"

    df = df_all.copy() if df_all is not None else pd.DataFrame()
    cols = [
        "signal_side", "signal_date", "symbol", "name", "板块", "signal_type", "model",
        "close", "RSI", "L2_trend", "L2_pump", "H4_RSI", "H4_FJ", "extra_info",
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = np.nan
    df = df[cols].copy()
    if not df.empty:
        df["signal_date"] = pd.to_datetime(df["signal_date"], errors="coerce").dt.date.astype(str)
        df = df.sort_values(["signal_side", "signal_date", "symbol"], ascending=[True, False, True])

    def fmt_num(x):
        try:
            if pd.isna(x):
                return ""
            return f"{float(x):.2f}"
        except Exception:
            return str(x)

    rows_html = []
    if df.empty:
        rows_html.append('<tr><td colspan="14" class="empty">No signals</td></tr>')
    else:
        for _, r in df.iterrows():
            side = str(r.get("signal_side", "")).upper()
            side_cls = "buy" if side == "BUY" else "sell"
            side_txt = "BUY" if side == "BUY" else "SELL"
            cells = [
                f'<td><span class="pill {side_cls}">{side_txt}</span></td>',
                f'<td>{r.get("signal_date", "")}</td>',
                f'<td class="sym">{r.get("symbol", "")}</td>',
                f'<td>{r.get("name", "")}</td>',
                f'<td>{r.get("板块", "")}</td>',
                f'<td>{r.get("signal_type", "")}</td>',
                f'<td>{r.get("model", "")}</td>',
                f'<td class="num">{fmt_num(r.get("close"))}</td>',
                f'<td class="num">{fmt_num(r.get("RSI"))}</td>',
                f'<td class="num">{fmt_num(r.get("L2_trend"))}</td>',
                f'<td class="num">{fmt_num(r.get("L2_pump"))}</td>',
                f'<td class="num">{fmt_num(r.get("H4_RSI"))}</td>',
                f'<td class="num">{fmt_num(r.get("H4_FJ"))}</td>',
                f'<td>{r.get("extra_info", "")}</td>',
            ]
            rows_html.append("<tr>" + "".join(cells) + "</tr>")

    html = """<!doctype html>
<html lang=\"zh-CN\">
<head>
<meta charset=\"utf-8\">
<title>XL Signal Dashboard</title>
<style>
body { margin: 0; font-family: -apple-system, BlinkMacSystemFont, \"Segoe UI\", sans-serif; background: #f5f7fb; color: #172033; }
.wrap { padding: 22px; }
h1 { margin: 0 0 4px; font-size: 22px; }
.sub { color: #667085; margin-bottom: 18px; }
table { width: 100%; border-collapse: collapse; background: #fff; border: 1px solid #d9e1ee; }
th { position: sticky; top: 0; background: #eef2f7; color: #344054; font-size: 12px; text-align: left; padding: 8px; border-bottom: 1px solid #d9e1ee; }
td { padding: 8px; border-bottom: 1px solid #edf1f6; font-size: 13px; vertical-align: top; }
tr:hover { background: #f8fbff; }
.sym { font-weight: 700; }
.num { text-align: right; font-variant-numeric: tabular-nums; }
.pill { display: inline-block; min-width: 50px; text-align: center; padding: 3px 7px; border-radius: 5px; color: white; font-size: 12px; font-weight: 700; }
.pill.buy { background: #00a884; }
.pill.short { background: #d92d20; }
.empty { text-align: center; color: #667085; padding: 30px; }
</style>
</head>
<body>
<div class=\"wrap\">
<h1>XL Signal Dashboard</h1>
<div class=\"sub\">Run: {run_time} · 第一买入点=低位首绿柱 · 二进宫买入点=回踩不破后再次绿柱 · 预警买入=4H BUY A/0出 · 正式买入=日线 BUY A/0出 · 预警卖出=4H 1出 · 正式卖出=日线 1出</div>
<table>
<thead><tr>
<th>方向</th><th>日期</th><th>代码</th><th>名称</th><th>板块</th><th>信号</th><th>模型</th>
<th>价格</th><th>RSI</th><th>L2趋势</th><th>L2泵</th><th>4H RSI</th><th>4H分金</th><th>说明</th>
</tr></thead>
<tbody>
{rows}
</tbody>
</table>
</div>
</body>
</html>
"""
    html = html.replace("{run_time}", run_dt.strftime('%Y-%m-%d %H:%M:%S')).replace("{rows}", "".join(rows_html))
    cur_path.write_text(html, encoding="utf-8")
    hist_path.write_text(html, encoding="utf-8")
    return cur_path, hist_path


def _recent_history_result_map(history_dir: Path, run_dt: datetime, max_days: int) -> dict[str, Path]:
    """
    最近 max_days 个交易日内已有结果文件的映射：
    {YYYY-MM-DD: /path/to/scan_result_...xlsx}
    """
    keep_dates = {d for d in pd.bdate_range(end=run_dt.date(), periods=max_days).date if d >= LIFECYCLE_START_DATE}
    pat = re.compile(r"^scan_result_(\d{8})_\d{6}\.xlsx$")
    out: dict[str, Path] = {}
    for p in sorted(history_dir.glob("scan_result_*.xlsx")):
        m = pat.match(p.name)
        if not m:
            continue
        try:
            dt = datetime.strptime(m.group(1), "%Y%m%d").date()
        except Exception:
            continue
        if dt not in keep_dates:
            continue
        out[dt.isoformat()] = p
    return out


def _get_catchup_signal_dates(history_dir: Path, run_dt: datetime, max_bdays: int = 5) -> list:
    """
    根据上一次成功扫描日期，自动补回遗漏的交易日信号。
    例如上次跑在 3/12，本次跑在 3/17，则补回 [3/13, 3/16, 3/17]。
    """
    pat = re.compile(r"^scan_result_(\d{8})_\d{6}\.xlsx$")
    last_run_date = None
    for p in sorted(history_dir.glob("scan_result_*.xlsx")):
        m = pat.match(p.name)
        if not m:
            continue
        try:
            d = datetime.strptime(m.group(1), "%Y%m%d").date()
        except Exception:
            continue
        if last_run_date is None or d > last_run_date:
            last_run_date = d

    today = run_dt.date()
    if last_run_date is None or last_run_date >= today:
        return [today]

    bdays = pd.bdate_range(start=last_run_date, end=today)
    dates = [x.date() for x in bdays][1:]
    if not dates:
        return [today]
    if len(dates) > max_bdays:
        dates = dates[-max_bdays:]
    return dates


def _get_forced_rescan_signal_dates(run_dt: datetime) -> list:
    """
    手动重建历史信号日期。
    设置 STOCK_ONECLICK_RESCAN_FROM=YYYY-MM-DD 时，保留该日期到本次运行日之间所有交易日信号。
    日常自动扫描不设置该变量，因此仍走 catchup 逻辑。
    """
    raw = str(os.environ.get("STOCK_ONECLICK_RESCAN_FROM", "") or "").strip()
    if not raw:
        return []
    start = pd.to_datetime(raw, errors="coerce")
    if pd.isna(start):
        print(f"⚠️ STOCK_ONECLICK_RESCAN_FROM 无法识别：{raw}，改用正常补回逻辑", flush=True)
        return []
    start_date = max(start.date(), LIFECYCLE_START_DATE)
    return [x.date() for x in pd.bdate_range(start=start_date, end=run_dt.date())]


def _build_followup_sheets(
    anchors: pd.DataFrame,
    run_dt: datetime,
    max_days: int = TRACK_MAX_DAYS,
    sector_map: dict[str, str] | None = None,
    sheet_prefix: str = "",
) -> tuple[dict[str, pd.DataFrame], dict[str, pd.DataFrame]]:
    """
    生成 {sheet_name: dataframe}，sheet_name = 信号日 YYYY-MM-DD。
    每个 sheet 包含：symbol, D0_date, D0_close, D1..D20 的 date/close/pct_vs_D0。
    """
    if anchors.empty:
        return {}, {}

    today = run_dt.date()
    symbols = sorted(anchors["symbol"].unique().tolist())
    close_cache: dict[str, pd.Series] = {}

    for sym in symbols:
        try:
            d = download_daily(sym, period="1y")
        except Exception:
            d = None
        if d is None or d.empty:
            continue
        s = d["Close"].copy()
        s.index = pd.to_datetime(s.index).date
        close_cache[sym] = s

    base_cols = ["symbol", "观海买点分", "板块", "D0_date", "D0_rule", "D0_close", "prior_14d_signal_dates"]
    sector_map = sector_map or {}

    out_rows: dict[str, list[dict]] = {}
    anchors = anchors.copy()
    anchors["signal_date"] = pd.to_datetime(anchors["signal_date"], errors="coerce").dt.date
    if "buy_score" not in anchors.columns:
        anchors["buy_score"] = np.nan
    anchors["buy_score"] = pd.to_numeric(anchors["buy_score"], errors="coerce")
    anchors = anchors.dropna(subset=["signal_date"])
    # 按 symbol 构建“活动周期锁定”：
    # - 首次触发作为 D0
    # - D0 后 max_days 交易日内重复触发不重置，只记录 retrigger_dates
    # - 超过 max_days 交易日后再次触发，开启新周期
    for sym in symbols:
        series = close_cache.get(sym)
        if series is None or series.empty:
            continue

        sym_anchor = anchors.loc[anchors["symbol"] == sym].copy()
        sig_dates = sorted(set(sym_anchor["signal_date"].tolist()))
        sig_dates = [d for d in sig_dates if d in series.index and d <= today]
        if not sig_dates:
            continue

        dates = list(series.index)
        pos = {d: i for i, d in enumerate(dates)}

        cycles: list[tuple[datetime.date, list[datetime.date]]] = []
        prior_by_signal_date: dict[datetime.date, list[datetime.date]] = {}
        active_anchor = None
        retriggers: list = []
        for d in sig_dates:
            if active_anchor is None:
                active_anchor = d
                retriggers = []
                prior_by_signal_date[d] = []
                continue
            if (pos[d] - pos[active_anchor]) > max_days:
                cycles.append((active_anchor, retriggers))
                active_anchor = d
                retriggers = []
                prior_by_signal_date[d] = []
            elif d != active_anchor:
                prior_by_signal_date[d] = [active_anchor] + retriggers.copy()
                retriggers.append(d)
        if active_anchor is not None:
            cycles.append((active_anchor, retriggers))

        def format_prior_signal_dates(prior_dates):
            parts = []
            for prior_date in prior_dates or []:
                score_hit = sym_anchor.loc[sym_anchor["signal_date"] == prior_date, "buy_score"]
                score = pd.to_numeric(score_hit, errors="coerce").max() if not score_hit.empty else np.nan
                if pd.notna(score):
                    score_txt = f"{float(score):.0f}" if float(score).is_integer() else f"{float(score):.1f}"
                    parts.append(f"{prior_date.isoformat()} ({score_txt})")
                else:
                    parts.append(prior_date.isoformat())
            return ", ".join(parts)

        def make_followup_row(sig_date, retrigs=None, prior_dates=None):
            retrigs = retrigs or []
            prior_dates = prior_dates or []
            base_loc = pos[sig_date]
            d0 = float(series.loc[sig_date])
            rule_hit = sym_anchor.loc[sym_anchor["signal_date"] == sig_date, "d0_rule"]
            d0_rule = str(rule_hit.iloc[0]).strip() if not rule_hit.empty else ""
            score_hit = sym_anchor.loc[sym_anchor["signal_date"] == sig_date, "buy_score"]
            buy_score = pd.to_numeric(score_hit, errors="coerce").max() if not score_hit.empty else np.nan

            latest_loc = base_loc
            for j in range(base_loc, len(dates)):
                if dates[j] <= today:
                    latest_loc = j
                else:
                    break
            max_available = latest_loc - base_loc
            if max_available < 0:
                return None
            if max_available > max_days:
                max_available = max_days

            row = {
                "symbol": sym,
                "观海买点分": round(float(buy_score), 1) if pd.notna(buy_score) else np.nan,
                "板块": sector_map.get(sym, "99 未分组"),
                "D0_date": sig_date.isoformat(),
                "D0_rule": d0_rule,
                "D0_close": round(d0, 4),
                "prior_14d_signal_dates": format_prior_signal_dates(prior_dates),
                "retrigger_dates": ", ".join([x.isoformat() for x in retrigs]) if retrigs else "",
            }
            for i in range(1, max_days + 1):
                if i <= max_available and (base_loc + i) < len(dates):
                    di_date = dates[base_loc + i]
                    di_close = float(series.iloc[base_loc + i])
                    di_pct = (di_close / d0 - 1.0) if d0 else np.nan
                    row[f"D{i}_date"] = di_date.isoformat()
                    row[f"D{i}_close"] = round(di_close, 4)
                    row[f"D{i}_pct_vs_D0"] = round(di_pct, 6) if pd.notna(di_pct) else np.nan
                else:
                    row[f"D{i}_date"] = np.nan
                    row[f"D{i}_close"] = np.nan
                    row[f"D{i}_pct_vs_D0"] = np.nan
            return row

        for sig_date, retrigs in cycles:
            row = make_followup_row(sig_date, retrigs=retrigs, prior_dates=[])
            if row is None:
                continue
            sheet_key = f"{sheet_prefix}{sig_date.isoformat()}"
            out_rows.setdefault(sheet_key, []).append(row)
            for retrig_date in retrigs:
                retrig_row = make_followup_row(
                    retrig_date,
                    retrigs=[],
                    prior_dates=prior_by_signal_date.get(retrig_date, []),
                )
                if retrig_row is None:
                    continue
                retrig_key = f"{sheet_prefix}{retrig_date.isoformat()}"
                out_rows.setdefault(retrig_key, []).append(retrig_row)

    out: dict[str, pd.DataFrame] = {}
    completed: dict[str, pd.DataFrame] = {}
    for sig_date, rows in out_rows.items():
        if not rows:
            continue
        df_sheet = pd.DataFrame(rows)
        if "观海买点分" in df_sheet.columns:
            df_sheet["_score_sort"] = pd.to_numeric(df_sheet["观海买点分"], errors="coerce").fillna(-1)
            df_sheet = df_sheet.sort_values(["_score_sort", "symbol"], ascending=[False, True]).drop(columns=["_score_sort"])
        else:
            df_sheet = df_sheet.sort_values("symbol")
        df_sheet = df_sheet.reset_index(drop=True)
        is_completed = f"D{max_days}_date" in df_sheet.columns and df_sheet[f"D{max_days}_date"].notna().all()

        ordered_cols = base_cols.copy()
        for i in range(1, max_days + 1):
            date_col = f"D{i}_date"
            close_col = f"D{i}_close"
            pct_col = f"D{i}_pct_vs_D0"
            if date_col not in df_sheet.columns:
                continue
            date_vals = df_sheet[date_col].dropna()
            if date_vals.empty:
                continue
            date_txt = str(date_vals.iloc[0])
            new_close_col = f"D{i}_{date_txt}"
            df_sheet = df_sheet.rename(columns={close_col: new_close_col})
            ordered_cols.extend([new_close_col, pct_col])

        if "prior_14d_signal_dates" not in ordered_cols and "prior_14d_signal_dates" in df_sheet.columns:
            ordered_cols.append("prior_14d_signal_dates")
        ordered_cols.append("retrigger_dates")
        df_sheet = df_sheet[[c for c in ordered_cols if c in df_sheet.columns]]
        if is_completed:
            completed[sig_date] = df_sheet
        else:
            out[sig_date] = df_sheet

    return out, completed


def _archive_completed_cycles(
    completed_sheets: dict[str, pd.DataFrame],
    completed_dir: Path,
    archive_prefix: str = "",
):
    """
    将已满 14 交易日的批次单独归档成文件。
    """
    if not completed_sheets:
        return []
    completed_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for sig_date, df_sheet in sorted(completed_sheets.items()):
        raw_date = sig_date.replace("SELL_", "")
        date_key = raw_date.replace("-", "")
        out_name = f"{date_key}共计14天数据.xlsx" if not archive_prefix else f"{archive_prefix}_{date_key}共计14天数据.xlsx"
        out_file = completed_dir / out_name
        # 已归档则不重复覆盖
        if out_file.exists():
            continue
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            sheet_name = sig_date[:31]
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_followup_sheet(writer.sheets[sheet_name])
            _append_sector_top5_stats(writer.sheets[sheet_name], df_sheet)
            _format_dates_and_autofit(writer.sheets[sheet_name])
        saved.append(out_file)
    return saved


def _style_followup_sheet(ws):
    """
    百分比列格式:
    - number format: 0.00%
    - 正数绿色，负数红色
    """
    if ws.max_row < 2:
        return
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    pct_cols = []
    for idx, h in enumerate(headers, start=1):
        htxt = str(h) if h is not None else ""
        if htxt.endswith("_pct_vs_D0"):
            pct_cols.append(idx)
    if not pct_cols:
        return

    pos_font = Font(color="FF008000")  # green
    neg_font = Font(color="FFFF0000")  # red

    for col in pct_cols:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            v = cell.value
            if v is None or (isinstance(v, float) and np.isnan(v)):
                continue
            if isinstance(v, (int, float)):
                cell.number_format = "0.00%"
                if v > 0:
                    cell.font = pos_font
                elif v < 0:
                    cell.font = neg_font


def _pct_text(value) -> str:
    try:
        if pd.isna(value):
            return "-"
        return f"{float(value):+.2%}"
    except Exception:
        return "-"


def _market_symbol_snapshot(symbol: str, name: str, xl: XunLongIndicator, as_of_date) -> dict:
    out = {"symbol": symbol, "name": name, "ok": False, "signals": pd.DataFrame()}
    try:
        daily = download_daily(symbol, period="1y")
    except Exception as exc:
        out["error"] = str(exc)
        return out
    if daily is None or daily.empty:
        out["error"] = "no daily data"
        return out

    daily = daily.copy()
    daily.index = pd.to_datetime(daily.index).date
    daily = daily[daily.index <= as_of_date]
    if daily.empty:
        out["error"] = "no daily data before as_of"
        return out

    close = pd.to_numeric(daily["Close"], errors="coerce")
    last_close = float(close.iloc[-1])
    prev_close = float(close.iloc[-2]) if len(close) >= 2 and pd.notna(close.iloc[-2]) else np.nan
    ma10 = float(close.rolling(10).mean().iloc[-1]) if len(close) >= 10 else np.nan
    ma20 = float(close.rolling(20).mean().iloc[-1]) if len(close) >= 20 else np.nan
    ma50 = float(close.rolling(50).mean().iloc[-1]) if len(close) >= 50 else np.nan
    pct_1d = (last_close / prev_close - 1.0) if prev_close and pd.notna(prev_close) else np.nan
    pct_5d = (last_close / float(close.iloc[-6]) - 1.0) if len(close) >= 6 and close.iloc[-6] else np.nan

    try:
        signals = scan_one_symbol(symbol, name, xl)
    except Exception:
        signals = pd.DataFrame()
    if signals is not None and not signals.empty:
        signals = signals.copy()
        signals["signal_date"] = pd.to_datetime(signals["signal_date"], errors="coerce").dt.date
        signals = signals[signals["signal_date"] <= as_of_date]
    else:
        signals = pd.DataFrame()

    out.update(
        {
            "ok": True,
            "last_close": last_close,
            "pct_1d": pct_1d,
            "pct_5d": pct_5d,
            "ma10": ma10,
            "ma20": ma20,
            "ma50": ma50,
            "below_ma10": pd.notna(ma10) and last_close < ma10,
            "below_ma20": pd.notna(ma20) and last_close < ma20,
            "below_ma50": pd.notna(ma50) and last_close < ma50,
            "signals": signals,
        }
    )
    return out


def _has_recent_signal(snapshot: dict, signal_type: str, max_bdays: int, as_of_date) -> bool:
    signals = snapshot.get("signals")
    if signals is None or signals.empty:
        return False
    hits = signals[signals["signal_type"].astype(str).eq(signal_type)].copy()
    if hits.empty:
        return False
    for d in hits["signal_date"].dropna().tolist():
        try:
            days = _business_days_between(d, as_of_date)
        except Exception:
            continue
        if 0 <= days <= max_bdays:
            return True
    return False


def build_market_context(run_dt: datetime) -> dict:
    as_of_date = run_dt.date()
    xl = XunLongIndicator()
    snapshots = {
        sym: _market_symbol_snapshot(sym, name, xl, as_of_date)
        for sym, name, _group in MARKET_CONTEXT_SYMBOLS
    }
    spx = snapshots.get("SPX", {})
    qqq = snapshots.get("QQQ", {})
    iwm = snapshots.get("IWM", {})
    rsp = snapshots.get("RSP", {})
    vix = snapshots.get("VIX", {})
    hyg = snapshots.get("HYG", {})
    smh = snapshots.get("SMH", {})
    xlu = snapshots.get("XLU", {})

    risk = 0
    reasons = []
    h4_notes = []

    for sym, label, weight in [("SPX", "SPX", 2), ("QQQ", "QQQ", 2), ("IWM", "IWM", 1)]:
        snap = snapshots.get(sym, {})
        if not snap.get("ok"):
            continue
        if _has_recent_signal(snap, "正式卖出", 3, as_of_date):
            risk += weight
            reasons.append(f"{label} 日线正式卖出")
        if _has_recent_signal(snap, "预警卖出", 2, as_of_date):
            risk += 1
            h4_notes.append(f"{label} 4H预警卖出")
        if snap.get("below_ma20"):
            risk += 1
            reasons.append(f"{label} 跌破MA20")

    if spx.get("ok") and pd.notna(spx.get("pct_1d", np.nan)) and spx["pct_1d"] <= -0.012:
        risk += 1
        reasons.append(f"SPX单日{_pct_text(spx['pct_1d'])}")
    if qqq.get("ok") and pd.notna(qqq.get("pct_1d", np.nan)) and qqq["pct_1d"] <= -0.018:
        risk += 1
        reasons.append(f"QQQ单日{_pct_text(qqq['pct_1d'])}")
    if vix.get("ok"):
        if pd.notna(vix.get("pct_1d", np.nan)) and vix["pct_1d"] >= 0.05:
            risk += 1
            reasons.append(f"VIX单日{_pct_text(vix['pct_1d'])}")
        if not vix.get("below_ma20", True):
            risk += 1
            reasons.append("VIX站上MA20")

    if rsp.get("ok"):
        if rsp.get("below_ma20"):
            risk += 1
            reasons.append("RSP跌破MA20")
        if spx.get("ok") and pd.notna(rsp.get("pct_5d", np.nan)) and pd.notna(spx.get("pct_5d", np.nan)):
            if rsp["pct_5d"] + 0.008 < spx["pct_5d"]:
                risk += 1
                reasons.append("RSP弱于SPX，市场宽度变差")

    if hyg.get("ok"):
        if hyg.get("below_ma20"):
            risk += 1
            reasons.append("HYG跌破MA20")
        if pd.notna(hyg.get("pct_5d", np.nan)) and hyg["pct_5d"] <= -0.01:
            risk += 1
            reasons.append(f"HYG 5日{_pct_text(hyg['pct_5d'])}")

    rotation_note = "轮动信号不足"
    if smh.get("ok") and xlu.get("ok"):
        smh_5d = smh.get("pct_5d", np.nan)
        xlu_5d = xlu.get("pct_5d", np.nan)
        if pd.notna(smh_5d) and pd.notna(xlu_5d):
            diff = xlu_5d - smh_5d
            if diff >= 0.02 and (smh.get("below_ma20") or smh_5d < 0):
                risk += 1
                reasons.append("XLU强于SMH，资金偏防守")
                rotation_note = f"防守强于进攻：XLU 5日{_pct_text(xlu_5d)} vs SMH 5日{_pct_text(smh_5d)}"
            elif smh_5d > xlu_5d:
                rotation_note = f"进攻强于防守：SMH 5日{_pct_text(smh_5d)} vs XLU 5日{_pct_text(xlu_5d)}"
            else:
                rotation_note = f"防守略强：XLU 5日{_pct_text(xlu_5d)} vs SMH 5日{_pct_text(smh_5d)}"

    if risk >= 6:
        state = "看跌/避险"
        suggestion = "买点整体降权；暂停追高，高beta和进攻板块只观察；优先现金流/防守/低beta。"
    elif risk >= 4:
        state = "风险升高"
        suggestion = "买点降权；只看高分和结构清晰标的；防守板块买点按避险轮动理解。"
    elif risk >= 2:
        state = "谨慎看涨"
        suggestion = "趋势未完全转空，但仓位和追涨要收敛；优先强趋势或防守现金流。"
    else:
        state = "强势/中性看涨"
        suggestion = "市场环境允许正常观察买点；高分进攻股优先级更高。"

    if not reasons:
        reasons.append("主要指数未触发明显日线风险")
    if not h4_notes:
        h4_notes.append("无明显4H卖出预警")

    parts = []
    for sym in ["SPX", "QQQ", "IWM", "RSP", "VIX", "HYG", "SMH", "XLU"]:
        snap = snapshots.get(sym, {})
        if snap.get("ok"):
            parts.append(f"{sym} {_pct_text(snap.get('pct_1d'))} / 5日{_pct_text(snap.get('pct_5d'))}")

    return {
        "state": state,
        "risk_score": risk,
        "as_of": as_of_date.isoformat(),
        "daily_reason": "；".join(reasons[:5]),
        "h4_note": "；".join(h4_notes[:4]),
        "rotation_note": rotation_note,
        "index_snapshot": "；".join(parts),
        "suggestion": suggestion,
    }


def _write_market_context_block(ws, start_row: int, market_context: dict | None) -> int:
    if not market_context:
        return start_row

    state = str(market_context.get("state", ""))
    if "看跌" in state:
        fill = PatternFill(fill_type="solid", start_color="FFF4CCCC", end_color="FFF4CCCC")
    elif "风险" in state:
        fill = PatternFill(fill_type="solid", start_color="FFFFE599", end_color="FFFFE599")
    elif "谨慎" in state:
        fill = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
    else:
        fill = PatternFill(fill_type="solid", start_color="FFD9EAD3", end_color="FFD9EAD3")

    rows = [
        ("市场环境", f"{state}（风险分 {market_context.get('risk_score', '-')}; 截至 {market_context.get('as_of', '')}）"),
        ("日线判断", market_context.get("daily_reason", "")),
        ("4H提示", market_context.get("h4_note", "")),
        ("轮动判断", market_context.get("rotation_note", "")),
        ("指数快照", market_context.get("index_snapshot", "")),
        ("策略提示", market_context.get("suggestion", "")),
    ]
    for ridx, (label, text) in enumerate(rows, start=start_row):
        ws.cell(row=ridx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=ridx, column=2, value=text)
        for cidx in range(1, 8):
            ws.cell(row=ridx, column=cidx).fill = fill
        ws.cell(row=ridx, column=2).alignment = Alignment(wrap_text=True)
    return start_row + len(rows) + 1


def _format_dates_and_autofit(ws, max_width: int = 42, data_row_count: int | None = None):
    """
    通用表格美化：
    - 冻结首行
    - 日期列格式 yyyy-mm-dd
    - 自动列宽（避免 #####/截断）
    """
    if ws.freeze_panes is None:
        ws.freeze_panes = "A2"

    # 日期列格式（只应用到真实数据区，避免污染底部统计区）
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    date_end_row = ws.max_row if data_row_count is None else max(1, 1 + int(data_row_count))
    for idx, h in enumerate(headers, start=1):
        htxt = str(h) if h is not None else ""
        if "date" in htxt.lower():
            for r in range(2, date_end_row + 1):
                cell = ws.cell(row=r, column=idx)
                if cell.value is None:
                    continue
                cell.number_format = "yyyy-mm-dd"

    preferred_width = {
        "name": 36,
        "板块": 34,
        "extra_info": 42,
    }

    # 自动列宽
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        header = ws.cell(row=1, column=col).value
        header_txt = str(header) if header is not None else ""
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        # 留白 + 上限，避免极长文本把表撑爆
        width = min(max(10, max_len + 2), max_width)
        if header_txt in preferred_width:
            width = max(width, preferred_width[header_txt])
        ws.column_dimensions[letter].width = width


def _style_summary_sheet(ws, summary_df: pd.DataFrame):
    """
    Summary 扩展样式：
    - 统计Top5板块（按出现只数降序）写到表格下方
    - Top5板块对应明细行浅蓝高亮
    """
    if summary_df.empty or "板块" not in summary_df.columns:
        return

    counts = (
        summary_df["板块"]
        .astype(str)
        .str.strip()
        .replace("", "99 未分组")
        .value_counts()
        .head(5)
    )
    if counts.empty:
        return

    top5 = set(counts.index.tolist())
    light_blue = PatternFill(fill_type="solid", start_color="FFDDEBF7", end_color="FFDDEBF7")

    # 高亮 Top5 板块对应行
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    sector_col = None
    for idx, h in enumerate(headers, start=1):
        if str(h) == "板块":
            sector_col = idx
            break
    if sector_col is not None:
        for r in range(2, ws.max_row + 1):
            v = str(ws.cell(row=r, column=sector_col).value or "").strip()
            if v in top5:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = light_blue

    # 底部统计
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value="板块Top5统计（按出现只数）")
    ws.cell(row=start_row + 1, column=1, value="排名")
    ws.cell(row=start_row + 1, column=2, value="板块")
    ws.cell(row=start_row + 1, column=3, value="股票只数")
    # 避免被 run_date 列格式影响统计区显示
    for rr in range(start_row, start_row + 8):
        ws.cell(row=rr, column=1).number_format = "General"
        ws.cell(row=rr, column=2).number_format = "General"
        ws.cell(row=rr, column=3).number_format = "General"
    for i, (sec, cnt) in enumerate(counts.items(), start=1):
        rr = start_row + 1 + i
        c_rank = ws.cell(row=rr, column=1, value=i)
        ws.cell(row=rr, column=2, value=sec)
        c_cnt = ws.cell(row=rr, column=3, value=int(cnt))
        c_rank.number_format = "0"
        c_cnt.number_format = "0"


def _highlight_top5_sector_rows(
    ws,
    start_row: int,
    end_row: int,
    sector_col_idx: int,
    top5: set[str],
    max_col: int,
):
    light_blue = PatternFill(fill_type="solid", start_color="FFDDEBF7", end_color="FFDDEBF7")
    for r in range(start_row, end_row + 1):
        v = str(ws.cell(row=r, column=sector_col_idx).value or "").strip()
        if v in top5:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = light_blue


def _is_priority_buy_rule(rule_text: str) -> bool:
    rule = str(rule_text or "")
    parts = [x.strip() for x in rule.split("|") if x.strip()]
    if len(parts) < 2:
        return False
    return (
        "正式买入" in rule
        and (
            "第一买入点" in rule
            or "预警买入" in rule
            or "二进宫买入点" in rule
        )
    )


def _highlight_priority_rule_rows(
    ws,
    start_row: int,
    end_row: int,
    rule_col_idx: int,
    max_col: int,
):
    priority_fill = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
    priority_font = Font(bold=True, color="FF9C5700")
    for r in range(start_row, end_row + 1):
        rule_cell = ws.cell(row=r, column=rule_col_idx)
        if not _is_priority_buy_rule(str(rule_cell.value or "")):
            continue
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = priority_fill
        rule_cell.font = priority_font


def _append_sector_top5_stats_at_row(ws, df: pd.DataFrame, start_row: int, title: str):
    if df.empty or "板块" not in df.columns:
        return start_row

    counts = (
        df["板块"]
        .astype(str)
        .str.strip()
        .replace("", "99 未分组")
        .value_counts()
        .head(5)
    )
    if counts.empty:
        return start_row

    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row + 1, column=1, value="排名")
    ws.cell(row=start_row + 1, column=2, value="板块")
    ws.cell(row=start_row + 1, column=3, value="股票只数")
    for rr in range(start_row, start_row + 8):
        ws.cell(row=rr, column=1).number_format = "General"
        ws.cell(row=rr, column=2).number_format = "General"
        ws.cell(row=rr, column=3).number_format = "General"

    for i, (sec, cnt) in enumerate(counts.items(), start=1):
        r = start_row + 1 + i
        ws.cell(row=r, column=1, value=i).number_format = "0"
        ws.cell(row=r, column=2, value=sec)
        ws.cell(row=r, column=3, value=int(cnt)).number_format = "0"
    return start_row + 1 + len(counts)


def _write_sector_top5_block(ws, counts: pd.Series, start_row: int, start_col: int, title: str) -> int:
    if counts is None or counts.empty:
        return start_row
    ws.cell(row=start_row, column=start_col, value=title)
    ws.cell(row=start_row + 1, column=start_col, value="排名")
    ws.cell(row=start_row + 1, column=start_col + 1, value="板块")
    ws.cell(row=start_row + 1, column=start_col + 2, value="股票只数")
    for rr in range(start_row, start_row + 8):
        for cc in range(start_col, start_col + 3):
            ws.cell(row=rr, column=cc).number_format = "General"
    for i, (sec, cnt) in enumerate(counts.items(), start=1):
        r = start_row + 1 + i
        ws.cell(row=r, column=start_col, value=i).number_format = "0"
        ws.cell(row=r, column=start_col + 1, value=sec)
        ws.cell(row=r, column=start_col + 2, value=int(cnt)).number_format = "0"
    return start_row + 1 + len(counts)


def _sector_top5_counts(df: pd.DataFrame) -> pd.Series:
    if df.empty or "板块" not in df.columns:
        return pd.Series(dtype=int)
    return (
        df["板块"]
        .astype(str)
        .str.strip()
        .replace("", "99 未分组")
        .value_counts()
        .head(5)
    )


def _append_buy_sector_top5_stats_at_row(ws, df: pd.DataFrame, start_row: int):
    all_counts = _sector_top5_counts(df)
    end_left = _write_sector_top5_block(ws, all_counts, start_row, 1, "买入板块Top5统计（按出现只数）")

    first_df = pd.DataFrame()
    if df is not None and not df.empty and "_first_seen_14d" in df.columns:
        first_df = df[df["_first_seen_14d"].fillna(False).astype(bool)].copy()
    elif df is not None and not df.empty and "prior_14d_signal_dates" in df.columns:
        prior = df["prior_14d_signal_dates"].fillna("").astype(str).str.strip()
        first_df = df[prior.eq("")].copy()
    first_counts = _sector_top5_counts(first_df)
    end_right = _write_sector_top5_block(
        ws,
        first_counts,
        start_row,
        5,
        "买入板块Top5统计（按出现只数）首次出现",
    )
    return max(end_left, end_right)


def _extract_latest_followup_close(row: pd.Series) -> tuple[object, object]:
    latest_date = np.nan
    latest_close = np.nan
    for col in row.index:
        if not str(col).startswith("D") or "_pct_vs_D0" in str(col):
            continue
        if col == "D0_close":
            continue
        val = row[col]
        if pd.isna(val):
            continue
        if "_" in str(col):
            latest_date = str(col).split("_", 1)[1]
            latest_close = val
    return latest_date, latest_close


def _build_active_section_summary(
    followup_sheets: dict[str, pd.DataFrame],
    df_run: pd.DataFrame,
    run_dt: datetime,
    signal_side: str,
    df_current_signals: pd.DataFrame | None = None,
) -> pd.DataFrame:
    if not followup_sheets:
        return pd.DataFrame()

    side = str(signal_side).upper()
    name_map = (
        df_run[["symbol", "name"]]
        .drop_duplicates(subset=["symbol"], keep="first")
        .assign(symbol=lambda x: x["symbol"].astype(str).str.strip().str.upper())
        .set_index("symbol")["name"]
        .to_dict()
    )

    current_signal_map: dict[tuple[str, str], list[tuple[object, str]]] = {}
    if df_current_signals is not None and not df_current_signals.empty:
        cur = df_current_signals.copy()
        if "signal_side" in cur.columns:
            cur = cur[cur["signal_side"].astype(str).str.upper() == side]
        if not cur.empty:
            cur["symbol"] = cur["symbol"].astype(str).str.strip().str.upper()
            cur["signal_type"] = cur.get("signal_type", "").fillna("").astype(str)
            cur["signal_date"] = pd.to_datetime(cur["signal_date"], errors="coerce").dt.date
            cur = cur.dropna(subset=["symbol", "signal_date"])
            grouped = (
                cur.sort_values(["signal_date", "signal_type"])
                .groupby(["symbol", "signal_date"], as_index=False)
                .agg(signal_type=("signal_type", lambda s: " | ".join([x for x in pd.unique(s) if x])))
            )
            for _, rr in grouped.iterrows():
                key = (str(rr["symbol"]).strip().upper(), side)
                current_signal_map.setdefault(key, []).append((rr["signal_date"], str(rr["signal_type"]).strip()))

    rows = []
    for _, sdf in sorted(followup_sheets.items()):
        if sdf is None or sdf.empty:
            continue
        for _, row in sdf.iterrows():
            symbol = str(row.get("symbol", "")).strip().upper()
            if not symbol:
                continue
            d0_date = row.get("D0_date", np.nan)
            d0_close = pd.to_numeric(row.get("D0_close"), errors="coerce")
            latest_date, latest_close = _extract_latest_followup_close(row)
            latest_close = pd.to_numeric(latest_close, errors="coerce")
            if pd.isna(latest_close):
                latest_close = d0_close
                latest_date = d0_date
            latest_pct = np.nan
            if pd.notna(d0_close) and d0_close:
                latest_pct = latest_close / d0_close - 1.0
            retrigger_dates = str(row.get("retrigger_dates", "") or "").strip()
            retrigger_list = [x.strip() for x in retrigger_dates.split(",") if x.strip()]
            retrigger_count = len(retrigger_list)

            cycle_hits = current_signal_map.get((symbol, side), [])
            current_cycle_hits = []
            for hit_date, hit_rule in cycle_hits:
                if str(hit_date) == str(d0_date) or str(hit_date) in retrigger_list:
                    current_cycle_hits.append((hit_date, hit_rule))
            current_cycle_hits = sorted(current_cycle_hits, key=lambda x: (x[0], x[1]))
            current_hit_desc = ", ".join([
                f"{hit_date.isoformat()}({hit_rule})" if hit_rule else hit_date.isoformat()
                for hit_date, hit_rule in current_cycle_hits
            ])

            extra_parts = []
            if pd.notna(d0_close) and pd.notna(latest_pct):
                extra_parts.append(f"D0={d0_close:.2f}")
                extra_parts.append(f"最新={latest_date}")
                extra_parts.append(f"相对D0={latest_pct:.2%}")
            if current_hit_desc:
                extra_parts.append(f"本次触发={current_hit_desc}")
            if retrigger_count:
                extra_parts.append(f"历史重复{retrigger_count}次")
                extra_parts.append(f"重复日期={retrigger_dates}")

            rows.append({
                "run_date": run_dt.date(),
                "run_time": run_dt.strftime("%H:%M:%S"),
                "symbol": symbol,
                "name": name_map.get(symbol, ""),
                "板块": row.get("板块", "99 未分组"),
                "signal_date": d0_date,
                "signal_type": "买入跟踪" if side == "BUY" else "卖出跟踪",
                "signal_side": side,
                "model": "FOLLOWUP_HIT" if current_hit_desc else "FOLLOWUP",
                "close": round(float(latest_close), 4) if pd.notna(latest_close) else np.nan,
                "volume": np.nan,
                "vol_ma20": np.nan,
                "L2_trend": np.nan,
                "L2_pump": np.nan,
                "RSI": np.nan,
                "rank120": np.nan,
                "extra_info": "; ".join(extra_parts),
                "_today_hit_rank": 0 if current_hit_desc else 1,
            })

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    out["signal_date"] = pd.to_datetime(out["signal_date"], errors="coerce").dt.date
    out = out.sort_values(["_today_hit_rank", "signal_date", "symbol"], ascending=[True, True, True]).reset_index(drop=True)
    if "_today_hit_rank" in out.columns:
        out = out.drop(columns=["_today_hit_rank"])
    return out



def _business_days_between(start_date, end_date) -> int:
    s = pd.to_datetime(start_date, errors="coerce")
    e = pd.to_datetime(end_date, errors="coerce")
    if pd.isna(s) or pd.isna(e):
        return 0
    return int(np.busday_count(s.date(), e.date()))


def _collect_lifecycle_signal_rows(history_dir: Path, current_df: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for p in sorted(history_dir.glob("scan_result_*.xlsx")):
        for side in ["BUY", "SELL"]:
            try:
                tmp = _read_signal_rows_from_result(p, side)
            except Exception:
                tmp = pd.DataFrame()
            if not tmp.empty:
                frames.append(tmp)
    if current_df is not None and not current_df.empty:
        frames.append(current_df.copy())
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    needed = [
        "symbol", "name", "板块", "signal_date", "signal_type", "signal_side",
        "close", "model", "extra_info", "Gann_1_date", "Gann_1_price",
    ]
    for c in needed:
        if c not in out.columns:
            out[c] = np.nan
    out["symbol"] = out["symbol"].astype(str).str.strip().str.upper()
    out["signal_type"] = out["signal_type"].astype(str).str.strip()
    out["signal_side"] = out["signal_side"].astype(str).str.strip().str.upper()
    out["signal_date"] = pd.to_datetime(out["signal_date"], errors="coerce").dt.date
    out["Gann_1_date"] = pd.to_datetime(out["Gann_1_date"], errors="coerce").dt.date
    out["Gann_1_price"] = pd.to_numeric(out["Gann_1_price"], errors="coerce")
    out = out[out["signal_type"].isin(["第一买入点", "二进宫买入点", "预警买入", "正式买入", "预警卖出", "正式卖出"])]
    out = out.dropna(subset=["symbol", "signal_date"])
    out = out[out["signal_date"] >= LIFECYCLE_START_DATE]
    out = out.sort_values(["symbol", "signal_date", "signal_type"]).drop_duplicates(
        subset=["symbol", "signal_date", "signal_type"], keep="last"
    )
    return out.reset_index(drop=True)


def _build_lifecycle_tables(history_dir: Path, current_df: pd.DataFrame, run_dt: datetime, df_run: pd.DataFrame, min_days: int = 14):
    sig = _collect_lifecycle_signal_rows(history_dir, current_df)
    if sig.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    meta = df_run[["symbol", "name", "group"]].drop_duplicates(subset=["symbol"], keep="first").copy()
    meta["symbol"] = meta["symbol"].astype(str).str.strip().str.upper()
    meta["板块_meta"] = meta["group"].apply(_normalize_sector_with_code)
    meta = meta.set_index("symbol")

    buy_history_rows = []
    buy_observation_rows = []
    sell_history_rows = []
    sell_observation_rows = []
    today = run_dt.date()

    for symbol, g in sig.groupby("symbol"):
        g = g.sort_values("signal_date")
        formal_buys = g[g["signal_type"] == "正式买入"]
        formal_sells = g[g["signal_type"] == "正式卖出"]
        early_buys = g[g["signal_type"].isin(["第一买入点", "二进宫买入点", "预警买入"])]
        buy_starts = formal_buys if not formal_buys.empty else early_buys

        for _, b in buy_starts.iterrows():
            bdate = b["signal_date"]
            future_formal_sells = formal_sells[formal_sells["signal_date"] > bdate]
            prior_warning_buy = early_buys[early_buys["signal_date"] <= bdate].tail(1)
            warning_buy_date = prior_warning_buy.iloc[0]["signal_date"] if not prior_warning_buy.empty else np.nan
            days = _business_days_between(bdate, today)
            base = {
                "symbol": symbol,
                "name": b.get("name") or (meta.loc[symbol, "name"] if symbol in meta.index else ""),
                "板块": b.get("板块") if pd.notna(b.get("板块")) else (meta.loc[symbol, "板块_meta"] if symbol in meta.index else "99 未分组"),
                "预警买入日期": warning_buy_date,
                "正式买入日期": bdate if b["signal_type"] == "正式买入" else np.nan,
                "买入跟踪起点": bdate,
                "买入价": b.get("close", np.nan),
                "已跟踪交易日": days,
            }
            if not future_formal_sells.empty:
                s = future_formal_sells.iloc[0]
                sdate = s["signal_date"]
                sell_close = pd.to_numeric(s.get("close", np.nan), errors="coerce")
                buy_close = pd.to_numeric(b.get("close", np.nan), errors="coerce")
                ret = sell_close / buy_close - 1.0 if pd.notna(sell_close) and pd.notna(buy_close) and buy_close else np.nan
                stage_days = _business_days_between(bdate, sdate)
                buy_history_rows.append({
                    **base,
                    "正式卖出日期": sdate,
                    "卖出价": sell_close,
                    "阶段交易日": stage_days,
                    "阶段收益率": ret,
                    "归档原因": f"正式卖出（{stage_days}个交易日内）" if stage_days <= min_days else "正式卖出",
                    "状态": "阶段结束",
                })
                continue
            if days >= min_days:
                buy_observation_rows.append({
                    **base,
                    "观察原因": f"超过{min_days}个交易日未出现正式卖出",
                    "状态": "观察中",
                })

        for _, s in formal_sells.iterrows():
            sdate = s["signal_date"]
            future_formal_buys = formal_buys[formal_buys["signal_date"] > sdate]
            future_warning_buys = g[g["signal_type"].isin(["第一买入点", "二进宫买入点", "预警买入"]) & (g["signal_date"] > sdate)]
            one_out_date = s.get("Gann_1_date")
            if pd.isna(one_out_date):
                one_out_date = sdate
            one_out_price = pd.to_numeric(s.get("Gann_1_price", np.nan), errors="coerce")
            if pd.isna(one_out_price):
                one_out_price = pd.to_numeric(s.get("close", np.nan), errors="coerce")
            sell_close = pd.to_numeric(s.get("close", np.nan), errors="coerce")
            days = _business_days_between(sdate, today)
            base = {
                "symbol": symbol,
                "name": s.get("name") or (meta.loc[symbol, "name"] if symbol in meta.index else ""),
                "板块": s.get("板块") if pd.notna(s.get("板块")) else (meta.loc[symbol, "板块_meta"] if symbol in meta.index else "99 未分组"),
                "1出日期": one_out_date,
                "1出价格": one_out_price,
                "正式卖出日期": sdate,
                "正式卖出价": sell_close,
                "卖出跟踪起点": sdate,
                "已跟踪交易日": days,
            }
            if not future_formal_buys.empty:
                b = future_formal_buys.iloc[0]
                bdate = b["signal_date"]
                warning_buy = future_warning_buys[future_warning_buys["signal_date"] <= bdate].head(1)
                warning_buy_date = warning_buy.iloc[0]["signal_date"] if not warning_buy.empty else np.nan
                buy_close = pd.to_numeric(b.get("close", np.nan), errors="coerce")
                stage_days = _business_days_between(sdate, bdate)
                sell_history_rows.append({
                    **base,
                    "下一次预警买入日期": warning_buy_date,
                    "下一次正式买入日期": bdate,
                    "下一次正式买入价": buy_close,
                    "阶段交易日": stage_days,
                    "归档原因": f"正式买入（{stage_days}个交易日内）" if stage_days <= min_days else "正式买入",
                    "状态": "卖出阶段结束",
                })
                continue
            if days >= min_days:
                warning_buy_date = future_warning_buys.iloc[0]["signal_date"] if not future_warning_buys.empty else np.nan
                sell_observation_rows.append({
                    **base,
                    "下一次预警买入日期": warning_buy_date,
                    "观察原因": f"超过{min_days}个交易日未出现正式买入",
                    "状态": "卖出观察中",
                })

    buy_history_df = pd.DataFrame(buy_history_rows)
    buy_observation_df = pd.DataFrame(buy_observation_rows)
    sell_history_df = pd.DataFrame(sell_history_rows)
    sell_observation_df = pd.DataFrame(sell_observation_rows)
    if not buy_history_df.empty:
        buy_history_df = buy_history_df.sort_values(["正式卖出日期", "symbol"], ascending=[False, True]).reset_index(drop=True)
    if not buy_observation_df.empty:
        buy_observation_df = buy_observation_df.sort_values(["买入跟踪起点", "symbol"], ascending=[True, True]).reset_index(drop=True)
    if not sell_history_df.empty:
        sell_history_df = sell_history_df.sort_values(["下一次正式买入日期", "symbol"], ascending=[False, True]).reset_index(drop=True)
    if not sell_observation_df.empty:
        sell_observation_df = sell_observation_df.sort_values(["卖出跟踪起点", "symbol"], ascending=[True, True]).reset_index(drop=True)
    return buy_observation_df, buy_history_df, sell_observation_df, sell_history_df


def _write_simple_table_sheet(writer, sheet_name: str, df: pd.DataFrame):
    out = df.copy() if df is not None else pd.DataFrame()
    if out.empty:
        out = pd.DataFrame({"message": ["No rows"]})
    out.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    ws = writer.sheets[sheet_name[:31]]
    _format_dates_and_autofit(ws)
    if "阶段收益率" in out.columns:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col = headers.index("阶段收益率") + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "0.00%"

def _write_summary_sheet(writer, buy_df: pd.DataFrame, sell_df: pd.DataFrame):
    ws = writer.book.create_sheet("Summary")
    writer.sheets["Summary"] = ws
    header_fill = PatternFill(fill_type="solid", start_color="FFD9EAD3", end_color="FFD9EAD3")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=12)

    current_row = 1

    def write_section(title: str, section_df: pd.DataFrame, stats_title: str):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=title).font = title_font
        header_row = current_row + 1
        section_cols = list(section_df.columns) if not section_df.empty else [
            "run_date", "run_time", "symbol", "name", "板块", "signal_date", "signal_type",
            "signal_side", "model", "close", "volume", "vol_ma20", "L2_trend", "L2_pump",
            "RSI", "rank120", "extra_info",
        ]
        for cidx, col_name in enumerate(section_cols, start=1):
            cell = ws.cell(row=header_row, column=cidx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        data_start = header_row + 1
        if section_df.empty:
            ws.cell(row=data_start, column=1, value="No signals")
            current_row = data_start + 2
            return

        for ridx, row in enumerate(section_df.itertuples(index=False), start=data_start):
            for cidx, val in enumerate(row, start=1):
                ws.cell(row=ridx, column=cidx, value=val)

        data_end = data_start + len(section_df) - 1
        try:
            sector_col_idx = section_cols.index("板块") + 1
        except ValueError:
            sector_col_idx = None
        if sector_col_idx is not None and "板块" in section_df.columns:
            top5 = set(
                section_df["板块"].astype(str).str.strip().replace("", "99 未分组").value_counts().head(5).index.tolist()
            )
            _highlight_top5_sector_rows(ws, data_start, data_end, sector_col_idx, top5, len(section_cols))

        current_row = data_end + 2
        current_row = _append_sector_top5_stats_at_row(ws, section_df, current_row, stats_title) + 2

    write_section("买入跟踪（14交易日内）", buy_df, "买入板块Top5统计（按出现只数）")
    write_section("卖出跟踪（14交易日内）", sell_df, "卖出板块Top5统计（按出现只数）")
    ws.freeze_panes = "A3"
    _format_dates_and_autofit(ws)


def _write_signal_snapshot_section(
    ws,
    start_row: int,
    title: str,
    section_df: pd.DataFrame,
    stats_title: str,
) -> int:
    header_fill = PatternFill(fill_type="solid", start_color="FFD9EAD3", end_color="FFD9EAD3")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=12)
    ws.cell(row=start_row, column=1, value=title).font = title_font

    header_row = start_row + 1
    section_cols = list(section_df.columns) if not section_df.empty else [
        "run_date", "run_time", "symbol", "name", "板块", "signal_date", "signal_type",
        "signal_side", "model", "close", "volume", "vol_ma20", "L2_trend", "L2_pump",
        "RSI", "rank120", "extra_info",
    ]
    for cidx, col_name in enumerate(section_cols, start=1):
        cell = ws.cell(row=header_row, column=cidx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        if str(col_name) == "观海买点分":
            cell.alignment = Alignment(horizontal="center")

    data_start = header_row + 1
    if section_df.empty:
        ws.cell(row=data_start, column=1, value="No signals")
        return data_start + 2

    for ridx, row in enumerate(section_df.itertuples(index=False), start=data_start):
        for cidx, val in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx, value=val)

    data_end = data_start + len(section_df) - 1
    try:
        sector_col_idx = section_cols.index("板块") + 1
    except ValueError:
        sector_col_idx = None
    if sector_col_idx is not None and "板块" in section_df.columns:
        top5 = set(
            section_df["板块"].astype(str).str.strip().replace("", "99 未分组").value_counts().head(5).index.tolist()
        )
        _highlight_top5_sector_rows(ws, data_start, data_end, sector_col_idx, top5, len(section_cols))

    next_row = data_end + 2
    return _append_sector_top5_stats_at_row(ws, section_df, next_row, stats_title) + 2


def _write_combined_snapshot_sheet(
    writer,
    sheet_name: str,
    buy_df: pd.DataFrame,
    sell_df: pd.DataFrame,
    market_context: dict | None = None,
):
    ws = writer.book.create_sheet(sheet_name[:31])
    writer.sheets[sheet_name[:31]] = ws
    row = 1
    row = _write_market_context_block(ws, row, market_context)
    row = _write_signal_snapshot_section(
        ws,
        row,
        f"{sheet_name} 买入信号快照",
        buy_df,
        "买入板块Top5统计（按出现只数）",
    )
    row = _write_signal_snapshot_section(
        ws,
        row,
        f"{sheet_name} 卖出信号快照",
        sell_df,
        "卖出板块Top5统计（按出现只数）",
    )
    ws.freeze_panes = "A9" if market_context else "A3"
    _format_dates_and_autofit(ws)


def _append_sector_top5_stats(ws, df: pd.DataFrame):
    """
    在当前 sheet 底部追加板块Top5统计（简版，不做高亮）。
    """
    if df.empty or "板块" not in df.columns:
        return
    counts = (
        df["板块"]
        .astype(str)
        .str.strip()
        .replace("", "99 未分组")
        .value_counts()
        .head(5)
    )
    if counts.empty:
        return

    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value="板块Top5统计（按出现只数）")
    ws.cell(row=start_row + 1, column=1, value="排名")
    ws.cell(row=start_row + 1, column=2, value="板块")
    ws.cell(row=start_row + 1, column=3, value="股票只数")
    for rr in range(start_row, start_row + 8):
        ws.cell(row=rr, column=1).number_format = "General"
        ws.cell(row=rr, column=2).number_format = "General"
        ws.cell(row=rr, column=3).number_format = "General"

    for i, (sec, cnt) in enumerate(counts.items(), start=1):
        r = start_row + 1 + i
        ws.cell(row=r, column=1, value=i).number_format = "0"
        ws.cell(row=r, column=2, value=sec)
        ws.cell(row=r, column=3, value=int(cnt)).number_format = "0"


def _write_followup_section(
    ws,
    start_row: int,
    title: str,
    section_df: pd.DataFrame,
    stats_title: str,
) -> int:
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill(fill_type="solid", start_color="FFD9EAD3", end_color="FFD9EAD3")
    header_font = Font(bold=True)
    pos_font = Font(color="FF008000")
    neg_font = Font(color="FFFF0000")

    ws.cell(row=start_row, column=1, value=title).font = title_font
    header_row = start_row + 1
    if not section_df.empty:
        section_cols = [c for c in section_df.columns if not str(c).startswith("_")]
        display_df = section_df[section_cols].copy()
    else:
        section_cols = ["symbol", "观海买点分", "板块", "D0_date", "D0_rule", "D0_close", "retrigger_dates"]
        display_df = section_df
    for cidx, col_name in enumerate(section_cols, start=1):
        cell = ws.cell(row=header_row, column=cidx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    data_start = header_row + 1
    if section_df.empty:
        ws.cell(row=data_start, column=1, value="No signals")
        return data_start + 2

    pct_cols = []
    score_col_idx = None
    rule_col_idx = None
    for cidx, col_name in enumerate(section_cols, start=1):
        if str(col_name).endswith("_pct_vs_D0"):
            pct_cols.append(cidx)
        if str(col_name) == "观海买点分":
            score_col_idx = cidx
        if str(col_name) == "D0_rule":
            rule_col_idx = cidx

    for ridx, row in enumerate(display_df.itertuples(index=False), start=data_start):
        for cidx, val in enumerate(row, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=val)
            if score_col_idx is not None and cidx == score_col_idx:
                cell.alignment = Alignment(horizontal="center")
                if isinstance(val, (int, float)) and not pd.isna(val):
                    cell.number_format = "0"
            if cidx in pct_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = "0.00%"
                if val > 0:
                    cell.font = pos_font
                elif val < 0:
                    cell.font = neg_font

    data_end = data_start + len(display_df) - 1
    try:
        sector_col_idx = section_cols.index("板块") + 1
    except ValueError:
        sector_col_idx = None
    if sector_col_idx is not None and "板块" in display_df.columns:
        top5 = set(
            display_df["板块"].astype(str).str.strip().replace("", "99 未分组").value_counts().head(5).index.tolist()
        )
        _highlight_top5_sector_rows(ws, data_start, data_end, sector_col_idx, top5, len(section_cols))

    if rule_col_idx is not None:
        _highlight_priority_rule_rows(ws, data_start, data_end, rule_col_idx, len(section_cols))

    next_row = data_end + 2
    if "买入板块Top5统计" in str(stats_title):
        return _append_buy_sector_top5_stats_at_row(ws, section_df, next_row) + 2
    return _append_sector_top5_stats_at_row(ws, section_df, next_row, stats_title) + 2


def _write_combined_followup_sheet(
    writer,
    sheet_name: str,
    buy_df: pd.DataFrame,
    sell_df: pd.DataFrame,
    market_context: dict | None = None,
):
    ws = writer.book.create_sheet(sheet_name[:31])
    writer.sheets[sheet_name[:31]] = ws
    row = 1
    row = _write_market_context_block(ws, row, market_context)
    row = _write_followup_section(
        ws,
        row,
        f"{sheet_name} 买入触发样本跟踪",
        buy_df,
        "买入板块Top5统计（按出现只数）",
    )
    row = _write_followup_section(
        ws,
        row,
        f"{sheet_name} 卖出触发样本跟踪",
        sell_df,
        "卖出板块Top5统计（按出现只数）",
    )
    ws.freeze_panes = "A9" if market_context else "A3"
    _format_dates_and_autofit(ws)


def export_tradingview_lists(df_signals: pd.DataFrame, run_dt: datetime, base_dir: Path, history_dir: Path):
    """
    生成 TradingView 可粘贴清单：
    - 当天文件：<project_root>/tv_daily_signals/tv_today_YYYY-MM-DD.txt
    - 当天最新快捷文件：<project_root>/tv_daily_signals/tv_today_latest.txt
    - 历史文件：history/tv_today_YYYYMMDD_HHMMSS.txt
    """
    date_str = run_dt.strftime("%Y-%m-%d")
    ts_str = run_dt.strftime("%Y%m%d_%H%M%S")
    # 与 history 同级目录
    daily_dir = history_dir.parent / "tv_daily_signals"
    daily_dir.mkdir(parents=True, exist_ok=True)
    today_path = daily_dir / f"tv_today_{date_str}.txt"
    latest_path = daily_dir / "tv_today_latest.txt"
    history_path = history_dir / f"tv_today_{ts_str}.txt"

    if df_signals.empty:
        content = "# No signals today\n"
    else:
        rows = (
            df_signals[["symbol", "exchange"]]
            .fillna("")
            .drop_duplicates(subset=["symbol"], keep="first")
        )
        lines = [build_tv_symbol(str(r["symbol"]).strip().upper(), str(r["exchange"])) for _, r in rows.iterrows()]
        lines = [x for x in lines if x]
        content = "\n".join(lines) + ("\n" if lines else "")

    today_path.write_text(content, encoding="utf-8")
    latest_path.write_text(content, encoding="utf-8")
    history_path.write_text(content, encoding="utf-8")
    return today_path, history_path


def _rule_text_for_tv(rule_text: str) -> str:
    return str(rule_text or "").replace("|", "+").strip()


def score_buy_signal_row(row: pd.Series) -> float:
    if str(row.get("signal_side", "")).upper() != "BUY":
        return np.nan
    score = 50.0
    signal_type = str(row.get("signal_type", ""))
    model = str(row.get("model", ""))
    if signal_type == "正式买入":
        score += 20
    elif signal_type == "预警买入":
        score += 14
    elif signal_type == "二进宫买入点":
        score += 16
    elif signal_type == "第一买入点":
        score += 10
    if "BUY_A" in model or "0出" in model:
        score += 8

    rank120 = pd.to_numeric(row.get("rank120", np.nan), errors="coerce")
    if pd.notna(rank120):
        if rank120 <= 0.25:
            score += 10
        elif rank120 <= 0.45:
            score += 7
        elif rank120 <= 0.65:
            score += 3
        elif rank120 >= 0.85:
            score -= 5

    rsi = pd.to_numeric(row.get("RSI", np.nan), errors="coerce")
    if pd.notna(rsi):
        if 42 <= rsi <= 65:
            score += 7
        elif 35 <= rsi < 42:
            score += 3
        elif rsi > 75:
            score -= 6

    l2 = pd.to_numeric(row.get("L2_trend", np.nan), errors="coerce")
    if pd.notna(l2) and l2 <= 35:
        score += 5
    h4_fj = pd.to_numeric(row.get("H4_FJ", np.nan), errors="coerce")
    if pd.notna(h4_fj) and h4_fj <= 55:
        score += 4
    h4_rsi = pd.to_numeric(row.get("H4_RSI", np.nan), errors="coerce")
    if pd.notna(h4_rsi) and h4_rsi >= 45:
        score += 3
    return round(max(0.0, min(100.0, score)), 1)


def score_sell_signal_row(row: pd.Series) -> float:
    """卖出分 (sell-conviction score), 0-100, for SELL rows only (NaN otherwise).

    Symmetric mirror of score_buy_signal_row: higher = a stronger / more
    actionable sell (top-confirmed, high in its range with room to fall,
    overbought-and-rolling-over). Added 2026-06; the original code scored buys
    only. Weights are a first pass — calibrate against forward returns with
    backtest_score.py (for a sell, a NEGATIVE forward return is the "right" one).
    """
    if str(row.get("signal_side", "")).upper() != "SELL":
        return np.nan
    score = 50.0
    signal_type = str(row.get("signal_type", ""))
    model = str(row.get("model", ""))
    if signal_type == "正式卖出":
        score += 20
    elif signal_type == "预警卖出":
        score += 14
    if "1出" in model or "SELL_1" in model:
        score += 8

    rank120 = pd.to_numeric(row.get("rank120", np.nan), errors="coerce")
    if pd.notna(rank120):
        if rank120 >= 0.85:      # high in its range -> most room to fall
            score += 10
        elif rank120 >= 0.65:
            score += 7
        elif rank120 >= 0.45:
            score += 3
        elif rank120 <= 0.15:    # already washed out -> little downside left
            score -= 5

    rsi = pd.to_numeric(row.get("RSI", np.nan), errors="coerce")
    if pd.notna(rsi):
        if 55 <= rsi <= 70:      # rolling over from strength -> prime short
            score += 7
        elif rsi > 70:           # overbought, but can stay bid -> smaller credit
            score += 4
        elif rsi < 35:           # already oversold -> bounce risk
            score -= 6

    l2 = pd.to_numeric(row.get("L2_trend", np.nan), errors="coerce")
    if pd.notna(l2) and l2 >= 60:        # topping after a strong up-swing
        score += 5
    h4_fj = pd.to_numeric(row.get("H4_FJ", np.nan), errors="coerce")
    if pd.notna(h4_fj) and h4_fj >= 60:  # 4H overheated
        score += 4
    h4_rsi = pd.to_numeric(row.get("H4_RSI", np.nan), errors="coerce")
    if pd.notna(h4_rsi) and h4_rsi <= 55:  # 4H momentum fading
        score += 3
    return round(max(0.0, min(100.0, score)), 1)


def export_tv_buy_signal_notes(
    buy_followup_df: pd.DataFrame,
    run_dt: datetime,
    df_run: pd.DataFrame,
    base_dir: Path,
    history_dir: Path,
):
    """
    导出当日买入触发样本：
    - tv_buy_today_latest.txt：纯 TradingView 导入版
    - tv_buy_today_notes_latest.txt：带备注版
    """
    date_str = run_dt.strftime("%Y-%m-%d")
    ts_str = run_dt.strftime("%Y%m%d_%H%M%S")
    out_dir = history_dir.parent / "tv_buy_signals"
    out_dir.mkdir(parents=True, exist_ok=True)

    pure_path = out_dir / f"tv_buy_today_{date_str}.txt"
    pure_latest = out_dir / "tv_buy_today_latest.txt"
    notes_path = out_dir / f"tv_buy_today_notes_{date_str}.txt"
    notes_latest = out_dir / "tv_buy_today_notes_latest.txt"
    pure_hist = history_dir / f"tv_buy_today_{ts_str}.txt"
    notes_hist = history_dir / f"tv_buy_today_notes_{ts_str}.txt"

    ex_map = {}
    if df_run is not None and not df_run.empty and {"symbol", "exchange"}.issubset(df_run.columns):
        ex_map = (
            df_run[["symbol", "exchange"]]
            .fillna("")
            .assign(symbol=lambda x: x["symbol"].astype(str).str.strip().str.upper())
            .drop_duplicates(subset=["symbol"], keep="first")
            .set_index("symbol")["exchange"]
            .to_dict()
        )

    if buy_followup_df is None or buy_followup_df.empty:
        pure_content = ""
        notes_content = "# No buy signals today\n"
    else:
        rows = buy_followup_df.copy()
        for col in ["symbol", "板块", "D0_date", "D0_rule", "D0_close", "观海买点分"]:
            if col not in rows.columns:
                rows[col] = ""
        rows["symbol"] = rows["symbol"].astype(str).str.strip().str.upper()
        rows["观海买点分"] = pd.to_numeric(rows["观海买点分"], errors="coerce")
        rows = rows[rows["symbol"] != ""].copy()
        rows["_score_sort"] = rows["观海买点分"].fillna(-1)
        rows = rows.sort_values(["_score_sort", "symbol"], ascending=[False, True]).drop_duplicates("symbol", keep="first")

        pure_lines = []
        note_lines = [
            f"# {date_str} 当日买入触发样本",
            "# 格式：TradingView代码 | 触发日期 | 观海买点分 | 触发规则 | 板块 | D0_close",
        ]
        for _, r in rows.iterrows():
            symbol = str(r["symbol"]).strip().upper()
            tv_symbol = build_tv_symbol(symbol, ex_map.get(symbol, ""))
            if not tv_symbol:
                continue
            pure_lines.append(tv_symbol)
            score = r.get("观海买点分", np.nan)
            score_txt = "" if pd.isna(score) else f"{float(score):.0f}"
            note_lines.append(
                f"{tv_symbol} | {r.get('D0_date', '')} | {score_txt} | "
                f"{_rule_text_for_tv(r.get('D0_rule', ''))} | {r.get('板块', '')} | {r.get('D0_close', '')}"
            )

        pure_content = "\n".join(pure_lines) + ("\n" if pure_lines else "")
        notes_content = "\n".join(note_lines) + "\n"

    for path, content in [
        (pure_path, pure_content),
        (pure_latest, pure_content),
        (pure_hist, pure_content),
        (notes_path, notes_content),
        (notes_latest, notes_content),
        (notes_hist, notes_content),
    ]:
        path.write_text(content, encoding="utf-8")

    return pure_path, notes_path, pure_latest, notes_latest


def export_tv_a_pool(df_meta: pd.DataFrame, run_dt: datetime, base_dir: Path, history_dir: Path):
    """
    生成固定 A 池清单：
    - 当前文件：tv_A_pool.txt
    - 运行快照：history/tv_A_pool_YYYYMMDD_HHMMSS.txt
    - 若内容变化：先把旧版备份到 history/tv_A_pool_prev_YYYYMMDD_HHMMSS.txt
    """
    ts_str = run_dt.strftime("%Y%m%d_%H%M%S")
    current_path = base_dir / "tv_A_pool.txt"
    snapshot_path = history_dir / f"tv_A_pool_{ts_str}.txt"
    prev_backup_path = None

    ex_map = {}
    if not df_meta.empty and "symbol" in df_meta.columns and "exchange" in df_meta.columns:
        for _, r in df_meta[["symbol", "exchange"]].fillna("").iterrows():
            sym = str(r["symbol"]).strip().upper()
            ex = str(r["exchange"]).strip()
            if sym and sym not in ex_map:
                ex_map[sym] = ex

    lines = [build_tv_symbol(sym, ex_map.get(sym, "")) for sym in A_POOL_SYMBOLS]
    content = "\n".join(lines) + "\n"

    if current_path.exists():
        old = current_path.read_text(encoding="utf-8")
        if old != content:
            prev_backup_path = history_dir / f"tv_A_pool_prev_{ts_str}.txt"
            prev_backup_path.write_text(old, encoding="utf-8")

    current_path.write_text(content, encoding="utf-8")
    snapshot_path.write_text(content, encoding="utf-8")
    return current_path, snapshot_path, prev_backup_path


def _safe_group_filename(group_name: str) -> str:
    bad = '\\/:*?"<>|'
    out = str(group_name or "未分组").strip()
    for ch in bad:
        out = out.replace(ch, "_")
    out = out.replace(" ", "_")
    return out or "未分组"


def _group_to_cn(group_name: str) -> str:
    """
    板块名称统一中文化：
    - 先保留用户已有中文分组
    - 对常见英文行业名做映射
    """
    g = str(group_name or "").strip()
    if not g:
        return "未分组"

    # 已是中文则直接返回
    if any("\u4e00" <= ch <= "\u9fff" for ch in g):
        return g

    mp = {
        "Technology": "科技",
        "Industrials": "工业",
        "Basic Materials": "基础材料",
        "Financial Services": "金融服务",
        "Consumer Cyclical": "可选消费",
        "Healthcare": "医疗健康",
        "Real Estate": "房地产",
        "Communication Services": "通信服务",
        "Energy": "能源",
        "Utilities": "公用事业",
        "Consumer Defensive": "必选消费",
        "nan": "未分组",
        "": "未分组",
    }
    return mp.get(g, g)


def _normalize_sector_with_code(group_name: str) -> str:
    """
    统一板块展示为“NN 板块名”，避免有的带编号有的不带编号。
    """
    raw = _group_to_cn(group_name)
    raw_text = str(raw or "").strip()
    numbered = re.match(r"^\s*(\d{1,2})(\.\d+)?\s+(.+?)\s*$", raw_text)
    if numbered:
        suffix = numbered.group(2) or ""
        return f"{int(numbered.group(1)):02d}{suffix} {numbered.group(3).strip()}"

    g = re.sub(r"^\s*\d{1,2}(?:\.\d+)?\s*", "", raw_text).strip()
    if not g:
        return "99 未分组"

    code_map = {
        "核心指数映射/超大盘核心": "00 核心指数映射/超大盘核心",
        "市场环境": "01 市场环境",
        "AI超级核心": "02 AI超级核心",
        "GPU / AI芯片": "03 GPU / AI芯片",
        "半导体主池": "03 GPU / AI芯片",
        "半导体设备 / EDA": "04 半导体设备 / EDA",
        "AI网络 / 光模块": "05 AI网络 / 光模块",
        "AI基建/光模块/电力链": "05 AI网络 / 光模块",
        "服务器 / 数据中心硬件": "06 服务器 / 数据中心硬件",
        "数据中心电力": "07 数据中心电力",
        "电力生产商": "07 数据中心电力",
        "核能": "08 核能",
        "储能": "09 储能",
        "电池储能": "09 储能",
        "太阳能": "10 太阳能",
        "AI软件 / SaaS": "11 AI软件 / SaaS",
        "AI软件 / SAAS": "11 AI软件 / SAAS",
        "AI软件/云/SaaS": "11 AI软件 / SaaS",
        "AI软件/云/SAAS": "11 AI软件 / SAAS",
        "科技": "11 AI软件 / SaaS",
        "网络安全": "12 网络安全",
        "材料 / 电网": "13 材料 / 电网",
        "基础材料": "13 材料 / 电网",
        "能源/原材料/周期": "13 材料 / 电网",
        "能源": "14 能源",
        "公用事业": "15 公用事业",
        "数据中心REIT": "16 数据中心REIT",
        "数据中心 REIT": "16 数据中心 REIT",
        "房地产": "16 数据中心REIT",
        "高弹性交易池": "17 高弹性交易池",
        "金融服务": "18 金融服务",
        "金融": "18 金融服务",
        "可选消费": "19 可选消费",
        "消费/平台/电商/出行": "19 可选消费",
        "通信服务": "20 通信服务",
        "数据中心": "21 数据中心",
        "工业": "21 工业",
        "工业/国防/运输/公用": "21 工业",
        "医疗": "22 医疗",
        "医疗健康": "22 医疗",
        "必选消费": "23 必选消费",
        "COWOS": "19 COWOS",
    }
    # 未命中时，保留原分类文字但统一加“98”前缀，避免无编号。
    return code_map.get(g, f"98 {g}")


def is_excluded_from_scan_group(group_name: str) -> bool:
    """
    不进入扫描池的系统/观察分组：
    - 列表最上方的 00/01 分组
    - 市场环境
    - 核心指数映射
    """
    raw = str(group_name or "").strip()
    normalized = _normalize_sector_with_code(raw)
    return (
        raw.startswith("00")
        or raw.startswith("01")
        or normalized.startswith("00")
        or normalized.startswith("01")
        or "市场环境" in raw
        or "市场环境" in normalized
        or "核心指数映射" in raw
        or "核心指数映射" in normalized
    )


def filter_scannable_universe(df_run: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = df_run.copy()
    if "group" not in rows.columns:
        rows["group"] = ""
    excluded_mask = rows["group"].apply(is_excluded_from_scan_group)
    return rows.loc[~excluded_mask].copy(), rows.loc[excluded_mask].copy()


def _drop_symbols(df: pd.DataFrame, symbols: set[str]) -> pd.DataFrame:
    if df is None or df.empty or not symbols or "symbol" not in df.columns:
        return df
    out = df.copy()
    mask = out["symbol"].astype(str).str.strip().str.upper().isin(symbols)
    return out.loc[~mask].reset_index(drop=True)


def export_tv_group_lists(df_run: pd.DataFrame, run_dt: datetime, base_dir: Path, history_dir: Path):
    """
    导出按分类分组的 TradingView 导入清单：
    - 当前目录：tv_groups/*.txt
    - 历史快照：history/tv_groups_YYYYMMDD_HHMMSS/*.txt
    说明：每个 txt 都可直接整段粘贴到 TradingView 对应列表中。
    """
    ts_str = run_dt.strftime("%Y%m%d_%H%M%S")
    groups_dir = base_dir / "tv_groups"
    hist_groups_dir = history_dir / f"tv_groups_{ts_str}"
    groups_dir.mkdir(parents=True, exist_ok=True)
    hist_groups_dir.mkdir(parents=True, exist_ok=True)

    rows = df_run.copy()
    for c in ["symbol", "exchange", "group"]:
        if c not in rows.columns:
            rows[c] = ""
    rows["symbol"] = rows["symbol"].astype(str).str.strip().str.upper()
    rows["exchange"] = rows["exchange"].astype(str).str.strip()
    rows["group"] = rows["group"].astype(str).str.strip().replace({"": "未分组"})
    rows = rows[rows["symbol"] != ""]
    rows = rows.drop_duplicates(subset=["symbol"], keep="first")

    # 按出现顺序保留分类顺序
    group_order = []
    seen = set()
    for g in rows["group"].tolist():
        if g not in seen:
            seen.add(g)
            group_order.append(g)

    index_lines = []
    file_count = 0
    for i, group_name in enumerate(group_order, start=1):
        sub = rows[rows["group"] == group_name]
        if sub.empty:
            continue
        lines = [
            build_tv_symbol(sym, ex)
            for sym, ex in zip(sub["symbol"].tolist(), sub["exchange"].tolist())
        ]
        lines = [x for x in lines if x]
        if not lines:
            continue
        content = "\n".join(lines) + "\n"
        safe_name = _safe_group_filename(group_name)
        fname = f"{i:02d}_{safe_name}.txt"
        cur_path = groups_dir / fname
        hist_path = hist_groups_dir / fname
        cur_path.write_text(content, encoding="utf-8")
        hist_path.write_text(content, encoding="utf-8")
        index_lines.append(f"{i:02d}. {group_name} ({len(lines)}) -> {fname}")
        file_count += 1

    index_text = "\n".join(index_lines) + ("\n" if index_lines else "")
    index_current = groups_dir / "_index.txt"
    index_history = hist_groups_dir / "_index.txt"
    index_current.write_text(index_text, encoding="utf-8")
    index_history.write_text(index_text, encoding="utf-8")
    return groups_dir, hist_groups_dir, index_current, index_history, file_count


def export_custom_watchlists_cn(run_dt: datetime, base_dir: Path, history_dir: Path):
    """
    按用户定义的中文分类导出 TradingView 清单：
    - 当前目录：tv_custom
    - 历史目录：history/tv_custom_YYYYMMDD_HHMMSS
    - 合并单文件：tv_watchlists_merged_cn.txt + history 同名时间戳备份
    """
    ts_str = run_dt.strftime("%Y%m%d_%H%M%S")
    cur_dir = base_dir / "tv_custom"
    hist_dir = history_dir / f"tv_custom_{ts_str}"
    cur_dir.mkdir(parents=True, exist_ok=True)
    hist_dir.mkdir(parents=True, exist_ok=True)

    index_lines = []
    merged_lines = ["目录"]

    for title, symbols in CUSTOM_WATCHLISTS_CN:
        index_lines.append(f"- {title}（{len(symbols)}）")

    merged_lines.extend(index_lines)
    merged_lines.append("")
    merged_lines.append("明细")

    file_count = 0
    for i, (title, symbols) in enumerate(CUSTOM_WATCHLISTS_CN, start=1):
        safe_title = _safe_group_filename(title)
        fname = f"{i:02d}_{safe_title}.txt"
        content = "\n".join(symbols) + "\n"
        (cur_dir / fname).write_text(content, encoding="utf-8")
        (hist_dir / fname).write_text(content, encoding="utf-8")
        file_count += 1

        merged_lines.append("")
        merged_lines.append(f"[{title}]")
        merged_lines.extend(symbols)

    index_text = "\n".join(index_lines) + "\n"
    (cur_dir / "_index.txt").write_text(index_text, encoding="utf-8")
    (hist_dir / "_index.txt").write_text(index_text, encoding="utf-8")

    merged_text = "\n".join(merged_lines) + "\n"
    merged_current = base_dir / "tv_watchlists_merged_cn.txt"
    merged_history = history_dir / f"tv_watchlists_merged_cn_{ts_str}.txt"
    merged_current.write_text(merged_text, encoding="utf-8")
    merged_history.write_text(merged_text, encoding="utf-8")

    return cur_dir, hist_dir, merged_current, merged_history, file_count


def export_full_scan_pool(df_run: pd.DataFrame, run_dt: datetime, base_dir: Path, history_dir: Path):
    """
    生成完整扫描池清单（用于 TradingView 快速导入）：
    - 当前文件：full_scan_pool.txt
    - 运行快照：history/full_scan_pool_YYYYMMDD_HHMMSS.txt
    - 若内容变化：先把旧版备份到 history/full_scan_pool_prev_YYYYMMDD_HHMMSS.txt
    """
    ts_str = run_dt.strftime("%Y%m%d_%H%M%S")
    current_path = base_dir / "full_scan_pool.txt"
    snapshot_path = history_dir / f"full_scan_pool_{ts_str}.txt"
    prev_backup_path = None

    if df_run.empty:
        content = "# Empty full scan pool\n"
    else:
        rows = (
            df_run[["symbol", "exchange"]]
            .fillna("")
            .drop_duplicates(subset=["symbol"], keep="first")
        )
        lines = [
            build_tv_symbol(str(r["symbol"]).strip().upper(), str(r["exchange"]))
            for _, r in rows.iterrows()
        ]
        lines = [x for x in lines if x]
        content = "\n".join(lines) + ("\n" if lines else "")

    if current_path.exists():
        old = current_path.read_text(encoding="utf-8")
        if old != content:
            prev_backup_path = history_dir / f"full_scan_pool_prev_{ts_str}.txt"
            prev_backup_path.write_text(old, encoding="utf-8")

    current_path.write_text(content, encoding="utf-8")
    snapshot_path.write_text(content, encoding="utf-8")
    return current_path, snapshot_path, prev_backup_path


def _fetch_yf_info(sym):
    """Fetch one symbol's yfinance .info, tolerant of failures. Returns (sym, dict)."""
    try:
        info = yf.Ticker(sym).info or {}
    except Exception:
        info = {}
    return sym, {
        "name": info.get("shortName") or info.get("longName") or "",
        "exchange": info.get("exchange", "") or "",
        "sector": info.get("sector", "") or "",
        "industry": info.get("industry", "") or "",
        "market_cap": info.get("marketCap", np.nan),
    }


def enrich_meta_with_yfinance(df_input, df_meta, force=None, max_workers=None):
    """
    用 yfinance 更新 Sheet2_Classified:
    - 新 symbol：加一行
    - 老 symbol：更新 name / market_cap / 行业等（不覆盖 group/note/enable）

    性能 / 稳定性优化：
    - 默认只联网拉取“新增”或“关键字段(name/exchange)缺失”的标的，避免每次运行都对
      全量 universe 调用极易被限流的 yfinance .info（这是之前最大的限流/拖慢来源）。
    - 设置 STOCK_ONECLICK_REFRESH_META=1 可强制刷新全部标的。
    - 需要拉取的标的用线程池并发获取（STOCK_ONECLICK_META_WORKERS，默认 8）。
    """
    meta_cols = [
        "symbol", "name", "exchange", "sector",
        "industry", "market_cap", "group", "note", "enable"
    ]
    if df_meta is None or df_meta.empty:
        df_meta = pd.DataFrame(columns=meta_cols)
    else:
        df_meta = df_meta.copy()

    if force is None:
        force = os.environ.get("STOCK_ONECLICK_REFRESH_META", "").strip() == "1"
    if max_workers is None:
        try:
            max_workers = int(os.environ.get("STOCK_ONECLICK_META_WORKERS", "8"))
        except ValueError:
            max_workers = 8

    symbols = df_input["symbol"].dropna().astype(str).str.strip().str.upper().unique().tolist()

    existing = {}
    if not df_meta.empty and "symbol" in df_meta.columns:
        for idx, r in df_meta.iterrows():
            existing[str(r["symbol"]).strip().upper()] = idx

    def _needs_fetch(sym):
        if force or sym not in existing:
            return True
        row = df_meta.loc[existing[sym]]
        name_ok = bool(str(row.get("name", "") or "").strip())
        exch_ok = bool(str(row.get("exchange", "") or "").strip())
        return not (name_ok and exch_ok)

    to_fetch = [s for s in symbols if _needs_fetch(s)]
    skipped = len(symbols) - len(to_fetch)
    print(
        f"元数据更新：共 {len(symbols)} 只，需联网 {len(to_fetch)} 只"
        f"（跳过已缓存 {skipped} 只{'，强制刷新' if force else ''}）",
        flush=True,
    )

    fetched = {}
    if to_fetch:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        workers = max(1, min(max_workers, len(to_fetch)))
        done = 0
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futures = [ex.submit(_fetch_yf_info, s) for s in to_fetch]
            for fut in as_completed(futures):
                done += 1
                try:
                    sym, info = fut.result()
                except Exception:
                    continue
                fetched[sym] = info
                if done % 25 == 0 or done == len(to_fetch):
                    print(f"[META {done}/{len(to_fetch)}]", flush=True)

    new_rows = []
    for sym in symbols:
        info = fetched.get(sym)
        if sym not in existing:
            info = info or {}
            sector = info.get("sector", "")
            new_rows.append({
                "symbol": sym,
                "name": info.get("name", ""),
                "exchange": info.get("exchange", ""),
                "sector": sector,
                "industry": info.get("industry", ""),
                "market_cap": info.get("market_cap", np.nan),
                "group": sector,   # 默认 group=sector，后面你可以手改
                "note": "",
                "enable": 1,
            })
        elif info is not None:
            # 老股票：只更新基础信息，不动 group/note/enable
            idx = existing[sym]
            if info.get("name"):
                df_meta.loc[idx, "name"] = info["name"]
            if info.get("exchange"):
                df_meta.loc[idx, "exchange"] = info["exchange"]
            if info.get("sector"):
                df_meta.loc[idx, "sector"] = info["sector"]
            if info.get("industry"):
                df_meta.loc[idx, "industry"] = info["industry"]
            mc = info.get("market_cap", np.nan)
            if mc is not None and not (isinstance(mc, float) and np.isnan(mc)):
                df_meta.loc[idx, "market_cap"] = mc

    if new_rows:
        df_meta = pd.concat([df_meta, pd.DataFrame(new_rows)], ignore_index=True)

    for col in meta_cols:
        if col not in df_meta.columns:
            df_meta[col] = np.nan

    df_meta["enable"] = pd.to_numeric(df_meta["enable"], errors="coerce").fillna(1).astype(int)

    return df_meta[meta_cols]


# ---- In-run bar cache + parallel prefetch -------------------------------------
# Keyed by (kind, yf_symbol, period). Used only by the nightly/CLI path: the
# dashboards replace download_daily / download_4h wholesale (monkey-patch), so
# they bypass this and keep their own provider-level cache. The cache removes the
# repeated fetches the scan + follow-up + market-context passes would otherwise
# do for the same symbol within a single run.
_BAR_CACHE: dict = {}
_BAR_CACHE_ENABLED = True


def clear_bar_cache():
    """Drop the in-run bar cache (call between independent runs in one process)."""
    _BAR_CACHE.clear()


def _fetch_daily_raw(symbol, period="1y"):
    yf_symbol = to_yfinance_symbol(symbol)
    # 用 Ticker().history() 而不是 yf.download()：download() 会写入 yfinance 的模块级
    # 全局状态 (shared._DFS)，多线程并发时不安全——并行预取下会把不同标的的列串台
    # （NUE 的请求曾返回 CEG/SMR 的数据，且根本不含 NUE），导致重复列 / 拿到错误标的。
    # Ticker().history() 是按实例的，线程安全，返回单层列、单只标的的数据。
    df = yf.Ticker(yf_symbol).history(period=period, interval="1d", auto_adjust=False)
    df = normalize_yf_df(df)
    if df.empty:
        return None
    # history() 的日线索引带时区(America/New_York)，download() 是无时区日期。
    # 去掉时区以保持与历史报告一致的日期语义（signal_date / 生命周期日期比较）。
    if getattr(df.index, "tz", None) is not None:
        df.index = df.index.tz_localize(None)
    return df[["Open", "High", "Low", "Close", "Volume"]].dropna()


def _fetch_4h_raw(symbol, period="90d"):
    yf_symbol = to_yfinance_symbol(symbol)
    # 同上：避免 yf.download() 的多线程串台问题，改用线程安全的 Ticker().history()。
    df = yf.Ticker(yf_symbol).history(period=period, interval="4h", auto_adjust=False)
    df = normalize_yf_df(df)
    if df.empty:
        return None
    # 旧的 download() 4H 索引是 UTC，这里把 history() 的 ET 索引转成 UTC 保持一致。
    if getattr(df.index, "tz", None) is not None:
        df.index = df.index.tz_convert("UTC")
    return df[["Open", "High", "Low", "Close", "Volume"]].dropna()


def _cached_bar(kind, symbol, period, fetcher):
    if not _BAR_CACHE_ENABLED:
        return fetcher(symbol, period)
    key = (kind, to_yfinance_symbol(symbol), period)
    if key not in _BAR_CACHE:
        _BAR_CACHE[key] = fetcher(symbol, period)
    cached = _BAR_CACHE[key]
    # Return a copy so downstream mutation can't poison the shared cache entry.
    return cached.copy() if cached is not None else None


def download_daily(symbol, period="1y"):
    return _cached_bar("daily", symbol, period, _fetch_daily_raw)


def download_4h(symbol, period="90d"):
    return _cached_bar("4h", symbol, period, _fetch_4h_raw)


def prefetch_bars(symbols, daily_period="1y", h4_period="90d", max_workers=None):
    """Warm the in-run bar cache in parallel so the serial scan / follow-up
    passes hit memory instead of the network. Best-effort: a failed fetch is
    cached as None and handled by the normal per-symbol skip logic.

    Concurrency override: STOCK_ONECLICK_DOWNLOAD_WORKERS (default 8)."""
    if not _BAR_CACHE_ENABLED:
        return
    syms = [str(s).strip().upper() for s in dict.fromkeys(symbols) if str(s).strip()]
    if not syms:
        return
    if max_workers is None:
        try:
            max_workers = int(os.environ.get("STOCK_ONECLICK_DOWNLOAD_WORKERS", "8"))
        except ValueError:
            max_workers = 8
    max_workers = max(1, min(max_workers, len(syms)))

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _warm(sym):
        # Distinct symbols → distinct cache keys, so threads never collide on a key.
        download_daily(sym, period=daily_period)
        download_4h(sym, period=h4_period)
        return sym

    print(f"并行预取行情：{len(syms)} 只 ×（日线+4H），并发 {max_workers} …", flush=True)
    done = 0
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(_warm, s): s for s in syms}
        for fut in as_completed(futures):
            done += 1
            if done % 25 == 0 or done == len(syms):
                print(f"  预取进度 {done}/{len(syms)}", flush=True)
            try:
                fut.result()
            except Exception:
                pass


# ================= 买点规则 =================

def add_V1_low_breakout(df):
    """SATS 模型：低位暴量突破"""
    df = df.copy()
    df["VolMA20"] = df["Volume"].rolling(20).mean()

    df["Close_min_120"] = df["Close"].rolling(120).min()
    df["Close_max_120"] = df["Close"].rolling(120).max()
    denom = df["Close_max_120"] - df["Close_min_120"]
    df["Rank120"] = (df["Close"] - df["Close_min_120"]) / denom.replace(0, np.nan)

    df["RecentHigh20"] = df["High"].rolling(20).max().shift(1)

    df["IsGreen"] = df["Close"] > df["Open"]
    df["CloseInUpperRange"] = (
        df["Close"] >= df["Low"] + 0.6 * (df["High"] - df["Low"])
    )

    df["VolSpike"] = df["Volume"] >= 3 * df["VolMA20"]
    df["LowRegionPrev"] = df["Rank120"].shift(1) <= 0.4
    df["Breakout20"] = df["Close"] > df["RecentHigh20"]

    df["V1_Buy"] = (
        df["IsGreen"]
        & df["CloseInUpperRange"]
        & df["VolSpike"]
        & df["LowRegionPrev"]
        & df["Breakout20"]
    )
    return df


def add_V2_daily_strong(df):
    """UBER 模型：日线走强（EMA8>、多头排列、RSI>=50）"""
    df = df.copy()
    # 低位过滤：若未提前计算 Rank120，则这里补一列
    if "Rank120" not in df.columns:
        close_min_120 = df["Close"].rolling(120).min()
        close_max_120 = df["Close"].rolling(120).max()
        denom = close_max_120 - close_min_120
        df["Rank120"] = (df["Close"] - close_min_120) / denom.replace(0, np.nan)

    df["EMA8_d"] = df["Close"].ewm(span=8, adjust=False).mean()
    df["SMA13_d"] = df["Close"].rolling(13).mean()
    df["SMA21_d"] = df["Close"].rolling(21).mean()

    delta = df["Close"].diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.rolling(14).mean()
    avg_loss = loss.rolling(14).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi_simple = 100 - 100 / (1 + rs)
    df["RSI_simple"] = rsi_simple

    cond_rank = df["Rank120"] <= V2_MAX_RANK120
    cond_price = df["Close"] > df["EMA8_d"]
    cond_ma = (df["EMA8_d"] > df["SMA13_d"]) & (df["SMA13_d"] > df["SMA21_d"])
    cond_rsi = df["RSI_simple"] >= 50

    df["DailyStrong"] = cond_rank & cond_price & cond_ma & cond_rsi
    return df


def add_buy_low_reset_4h_green(df):
    """
    低位修复 + 4H绿柱：
    - 日线分金近期打到低位（接近 0）
    - 当前分金重新抬升，近似“转黄柱”
    - 4H 出现首个绿柱确认
    - 日线至少止跌，避免还在单边下杀
    """
    df = df.copy()

    fj = pd.to_numeric(df.get("FJ_value"), errors="coerce")
    close = pd.to_numeric(df["Close"], errors="coerce")
    open_ = pd.to_numeric(df["Open"], errors="coerce")
    ema8 = pd.to_numeric(df.get("ema8"), errors="coerce")
    rank120 = pd.to_numeric(df.get("Rank120"), errors="coerce")

    fj_recent_low = fj.rolling(8, min_periods=1).min() <= 10
    fj_turn_up = (fj > fj.shift(1)) & (fj.shift(1) >= fj.shift(2))
    fj_yellow_zone = (fj >= 3) & (fj <= 45)
    fj_repair = fj_recent_low & fj_turn_up & fj_yellow_zone

    bfg = df.get("B_firstGreen_daily", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    c4_ok = df.get("C_ok_4h_daily", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    h4_confirm = bfg | c4_ok.rolling(2, min_periods=1).max().astype(bool)

    daily_stabilized = ((close >= ema8) | (close > close.shift(1))) & (close >= open_)
    low_context = rank120.fillna(1.0) <= 0.75

    df["BUY_low_reset_4h_green"] = low_context & fj_repair & h4_confirm & daily_stabilized
    return df


def add_buy_low_reset_confirmed(df):
    """
    低位修复买点的确认版：
    旧版只要分金低位修复 + 4H绿柱就提示，容易在下跌中太早出现。
    新版要求日线已经重新站回 EMA8、EMA8 抬升、RSI 回到 45 以上，
    且成交量不低于 20 日均量的 80%。
    """
    df = df.copy()
    base = df.get("BUY_low_reset_4h_green", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    close = pd.to_numeric(df["Close"], errors="coerce")
    open_ = pd.to_numeric(df["Open"], errors="coerce")
    ema8 = pd.to_numeric(df.get("ema8"), errors="coerce")
    rsi = pd.to_numeric(df.get("RSI"), errors="coerce")
    vol = pd.to_numeric(df["Volume"], errors="coerce")
    vol_ma20 = pd.to_numeric(df.get("VolMA20"), errors="coerce")

    price_confirmed = (close > ema8) & (ema8 > ema8.shift(1)) & ((close > open_) | (close > close.shift(1)))
    momentum_confirmed = (rsi >= 45) & (rsi <= 72)
    volume_confirmed = vol >= 0.8 * vol_ma20
    not_too_extended = close <= 1.12 * ema8

    df["BUY_low_reset_confirmed"] = (
        base
        & price_confirmed
        & momentum_confirmed
        & volume_confirmed
        & not_too_extended
    )
    return df


def add_low_start_buy_points(df):
    """
    用户偏好的低位启动器：
    - 第一买入点：低位首根日线绿柱/启动柱，只做预备观察。
    - 二进宫买入点：首绿后有真实回踩、结构未明显跌破，再次绿柱确认。
    """
    df = df.copy()
    close = pd.to_numeric(df["Close"], errors="coerce")
    high = pd.to_numeric(df["High"], errors="coerce")
    low = pd.to_numeric(df["Low"], errors="coerce")
    open_ = pd.to_numeric(df["Open"], errors="coerce")
    rank120 = pd.to_numeric(df.get("Rank120"), errors="coerce").fillna(1.0)
    l2_trend = pd.to_numeric(df.get("L2_trend"), errors="coerce").fillna(100.0)
    fj = pd.to_numeric(df.get("FJ_value"), errors="coerce").fillna(100.0)

    d1 = (close + low + high) / 3.0
    d2 = d1.ewm(span=6, adjust=False).mean()
    d3 = d2.ewm(span=5, adjust=False).mean()
    green_sig = (d2 > d3) & (d2.shift(1) <= d3.shift(1))

    low_context = (rank120 <= 0.45) | (l2_trend <= 30) | (fj <= 45)
    stabilized = (close >= open_) | (close > close.shift(1))
    valid_green = green_sig & low_context & stabilized

    first_vals = np.zeros(len(df), dtype=bool)
    second_vals = np.zeros(len(df), dtype=bool)
    prev_green_loc = None
    max_gap = 15
    min_pullback_pct = 0.025
    structure_break_allowance = 0.97

    for i, is_green in enumerate(valid_green.fillna(False).to_numpy()):
        if not is_green:
            continue

        is_second = False
        if prev_green_loc is not None:
            gap = i - prev_green_loc
            if 2 <= gap <= max_gap:
                prev_close = float(close.iloc[prev_green_loc])
                prev_low = float(low.iloc[prev_green_loc])
                pullback_low = float(low.iloc[prev_green_loc + 1 : i + 1].min())
                pullback_depth = (prev_close - pullback_low) / prev_close if prev_close else 0.0
                held_structure = pullback_low >= prev_low * structure_break_allowance
                is_second = pullback_depth >= min_pullback_pct and held_structure

        if is_second:
            second_vals[i] = True
        else:
            first_vals[i] = True
        prev_green_loc = i

    df["LOW_START_GREEN"] = valid_green
    df["LOW_START_FIRST_BUY"] = pd.Series(first_vals, index=df.index)
    df["LOW_START_SECOND_BUY"] = pd.Series(second_vals, index=df.index)
    return df


# ================= 扫描一个股票 =================

def scan_one_symbol(sym, name, xl: XunLongIndicator, *, daily_fetcher=None, h4_fetcher=None):
    # Data access is injectable: dashboards / tests / alternate providers can
    # pass their own fetchers. Defaults resolve to the module-level downloaders
    # at call time, so existing monkey-patching of scan.download_daily /
    # scan.download_4h keeps working unchanged.
    _dl = daily_fetcher or download_daily
    _h4 = h4_fetcher or download_4h
    df_d = _dl(sym, period="1y")
    if df_d is None or len(df_d) < 150:
        print(f"[{sym}] 日线数据太短/获取失败，跳过")
        return pd.DataFrame()

    df_4h = _h4(sym, period="90d")

    # 寻龙诀全套
    df_xl = xl.compute(df_d, df_4h)

    # V1 & V2
    df_v1 = add_V1_low_breakout(df_xl)
    df_v2 = add_V2_daily_strong(df_xl)
    df_v2_reset = add_buy_low_reset_4h_green(df_xl)
    df_v2_reset_confirmed = add_buy_low_reset_confirmed(df_v2_reset)
    df_low_start = add_low_start_buy_points(df_v2_reset_confirmed)

    df_all = df_xl.join(df_v1[["V1_Buy"]]).join(
        df_v2[["EMA8_d", "SMA13_d", "SMA21_d", "RSI_simple", "DailyStrong"]]
    ).join(
        df_v2_reset_confirmed[["BUY_low_reset_4h_green", "BUY_low_reset_confirmed"]]
    ).join(
        df_low_start[["LOW_START_GREEN", "LOW_START_FIRST_BUY", "LOW_START_SECOND_BUY"]]
    )

    rows = []
    recent_window = max(V1_LOOKBACK_DAYS, V2_LOOKBACK_DAYS, GANN_LOOKBACK_DAYS)
    recent_index = df_all.tail(recent_window).index
    daily_recent_mask = df_all.index.isin(recent_index)

    def append_signal(idx, row, signal_type, signal_side, model, extra_info):
        gann_1_date = pd.to_datetime(row.get("Gann_1_date", pd.NaT), errors="coerce")
        gann_1_date = gann_1_date.date() if pd.notna(gann_1_date) else np.nan
        rows.append({
            "symbol": sym,
            "name": name,
            "signal_date": idx.date(),
            "signal_type": signal_type,
            "signal_side": signal_side,
            "model": model,
            "close": row["Close"],
            "volume": row["Volume"],
            "vol_ma20": row.get("VolMA20", np.nan),
            "L2_trend": row["L2_trend"],
            "L2_pump": row["L2_pump"],
            "RSI": row["RSI"],
            "rank120": row.get("Rank120", np.nan),
            "H4_RSI": row.get("H4_RSI_last", np.nan),
            "H4_FJ": row.get("H4_FJ_last", np.nan),
            "H4_0_birth": row.get("H4_Gann_0_birth_daily", False),
            "H4_1_birth": row.get("H4_Gann_1_birth_daily", False),
            "Gann_1_date": gann_1_date,
            "Gann_1_price": row.get("Gann_1", np.nan),
            "extra_info": extra_info,
        })

    # ---- 新版买卖策略：只保留预警/正式四类 ----
    # 低位启动器：第一买入点 / 二进宫买入点
    if "LOW_START_FIRST_BUY" in df_all.columns:
        recent = df_all[df_all["LOW_START_FIRST_BUY"].fillna(False).astype(bool) & daily_recent_mask]
        for idx, row in recent.iterrows():
            append_signal(
                idx, row, "第一买入点", "BUY", "LOW_START_FIRST_GREEN",
                f"低位首绿柱; Rank120={row.get('Rank120', np.nan):.2f}; L2={row.get('L2_trend', np.nan):.2f}; 分金={row.get('FJ_value', np.nan):.2f}; 只做预备观察"
            )

    if "LOW_START_SECOND_BUY" in df_all.columns:
        recent = df_all[df_all["LOW_START_SECOND_BUY"].fillna(False).astype(bool) & daily_recent_mask]
        for idx, row in recent.iterrows():
            append_signal(
                idx, row, "二进宫买入点", "BUY", "LOW_START_SECOND_GREEN",
                f"首绿后回踩不破结构，再次绿柱确认; Rank120={row.get('Rank120', np.nan):.2f}; L2={row.get('L2_trend', np.nan):.2f}; 分金={row.get('FJ_value', np.nan):.2f}"
            )

    # 预警买入：4H BUY A / 4H 0出
    if "H4_Gann_0_birth_daily" in df_all.columns:
        recent = df_all[df_all["H4_Gann_0_birth_daily"].fillna(False).astype(bool) & daily_recent_mask]
        for idx, row in recent.iterrows():
            append_signal(
                idx, row, "预警买入", "BUY", "H4_BUY_A_0出",
                f"4H 0出; 4H_RSI={row.get('H4_RSI_last', np.nan):.2f}; 4H分金={row.get('H4_FJ_last', np.nan):.2f}"
            )

    # 正式买入：日线 BUY A / 日线 0出
    if "Gann_BUY_A" in df_all.columns:
        recent = df_all[df_all["Gann_BUY_A"].fillna(False).astype(bool) & daily_recent_mask]
        for idx, row in recent.iterrows():
            append_signal(
                idx, row, "正式买入", "BUY", "D1_BUY_A_0出",
                f"日线0出; Gann0={row.get('Gann_0', np.nan):.2f}; Gann1预览={row.get('Gann_1', np.nan):.2f}; 段涨幅={row.get('Gann_gain_pct', np.nan):.2%}"
            )

    # 预警卖出：4H 1出
    if "H4_Gann_1_birth_daily" in df_all.columns:
        recent = df_all[df_all["H4_Gann_1_birth_daily"].fillna(False).astype(bool) & daily_recent_mask]
        for idx, row in recent.iterrows():
            append_signal(
                idx, row, "预警卖出", "SELL", "H4_SELL_1出",
                f"4H 1出; 4H_RSI={row.get('H4_RSI_last', np.nan):.2f}; 4H分金={row.get('H4_FJ_last', np.nan):.2f}"
            )

    # 正式卖出：日线 1出
    if "Gann_SELL_1_confirmed" in df_all.columns:
        recent = df_all[df_all["Gann_SELL_1_confirmed"].fillna(False).astype(bool) & daily_recent_mask]
        for idx, row in recent.iterrows():
            one_date = pd.to_datetime(row.get("Gann_1_date", pd.NaT), errors="coerce")
            one_date_text = one_date.date().isoformat() if pd.notna(one_date) else ""
            append_signal(
                idx, row, "正式卖出", "SELL", "D1_SELL_1出",
                f"日线1出确认; 1出日期={one_date_text}; 1出价格={row.get('Gann_1', np.nan):.2f}; Gann0={row.get('Gann_0', np.nan):.2f}; 段涨幅={row.get('Gann_gain_pct', np.nan):.2%}"
            )

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows)


# ================= 主流程 =================

def main():
    if not INPUT_FILE.exists():
        print(f"找不到 {INPUT_FILE.name}，请确认文件在脚本同一目录。")
        return

    # 1) 读取输入 & meta
    df_input, df_meta = load_input_and_meta(INPUT_FILE)
    if df_input.empty:
        print("Sheet1_Input 里没有 symbol，先在第一列写几只股票代码（比如 UBER, NVDA）。")
        return

    # 2) 更新 meta 信息
    print("阶段 1/3：更新股票元数据...", flush=True)
    df_meta_new = enrich_meta_with_yfinance(df_input, df_meta)

    # 3) 写回 Excel（两个 sheet）
    print("阶段 2/3：写回输入与分类表...", flush=True)
    try:
        with pd.ExcelWriter(INPUT_FILE, engine="openpyxl") as writer:
            df_input.to_excel(writer, sheet_name="Sheet1_Input", index=False)
            df_meta_new.to_excel(writer, sheet_name="Sheet2_Classified", index=False)
    except PermissionError as exc:
        print(f"⚠️ 写回输入表失败，继续使用内存中的元数据扫描：{exc}", flush=True)

    # 4) 先导出固定 A 池 TradingView 清单（保留历史）
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    run_dt = datetime.now()
    tv_a_path, tv_a_hist_path, tv_a_prev_path = export_tv_a_pool(df_meta_new, run_dt, EXPORT_DIR, HISTORY_DIR)
    print(f"✅ TV固定A池：{tv_a_path}")
    print(f"✅ TV固定A池快照：{tv_a_hist_path}")
    if tv_a_prev_path is not None:
        print(f"✅ TV固定A池旧版备份：{tv_a_prev_path}")

    # 用户自定义中文分类清单（当前 + 历史 + 合并单文件）
    custom_dir, custom_hist_dir, custom_merged_cur, custom_merged_hist, custom_count = export_custom_watchlists_cn(
        run_dt, EXPORT_DIR, HISTORY_DIR
    )
    print(f"✅ TV中文分类目录：{custom_dir}（{custom_count}个分类）")
    print(f"✅ TV中文分类历史：{custom_hist_dir}")
    print(f"✅ TV中文合并清单：{custom_merged_cur}")
    print(f"✅ TV中文合并历史：{custom_merged_hist}")

    # 5) 取 enable=1 的股票跑扫描
    print("阶段 3/4：执行扫描...", flush=True)
    df_enabled = df_meta_new[df_meta_new["enable"] == 1]
    df_run, df_excluded = filter_scannable_universe(df_enabled)
    excluded_symbols = set()
    if not df_excluded.empty and "symbol" in df_excluded.columns:
        excluded_symbols = set(df_excluded["symbol"].astype(str).str.strip().str.upper())
    if not df_excluded.empty:
        skipped_groups = ", ".join(
            df_excluded["group"].fillna("").astype(str).replace("", "未分组").drop_duplicates().tolist()
        )
        print(f"跳过不扫描分组：{skipped_groups}（{len(df_excluded)} 只）", flush=True)
    if df_run.empty:
        print("过滤市场环境/顶部系统分组后，没有可扫描的标的。")
        return

    # 完整扫描池（当前 + 历史）
    full_pool_path, full_pool_hist_path, full_pool_prev_path = export_full_scan_pool(df_run, run_dt, EXPORT_DIR, HISTORY_DIR)
    print(f"✅ 全量扫描池：{full_pool_path}")
    print(f"✅ 全量扫描池快照：{full_pool_hist_path}")
    if full_pool_prev_path is not None:
        print(f"✅ 全量扫描池旧版备份：{full_pool_prev_path}")

    # 分类导入清单（当前 + 历史）
    tv_groups_dir, tv_groups_hist_dir, tv_groups_index_cur, tv_groups_index_hist, tv_group_count = export_tv_group_lists(
        df_run, run_dt, EXPORT_DIR, HISTORY_DIR
    )
    print(f"✅ TV分类清单目录：{tv_groups_dir}（{tv_group_count}个分类）")
    print(f"✅ TV分类清单索引：{tv_groups_index_cur}")
    print(f"✅ TV分类清单历史：{tv_groups_hist_dir}")

    # symbol -> 板块（统一编号中文）
    sector_map_all = (
        df_run[["symbol", "group"]]
        .drop_duplicates(subset=["symbol"], keep="first")
        .assign(板块=lambda x: x["group"].apply(_normalize_sector_with_code))
        .set_index("symbol")["板块"]
        .to_dict()
    )

    # 并行预取行情，热身 in-run 缓存，避免“扫描 + 追踪”两遍串行下载
    # （之前最大的串行网络开销）。失败的标的会缓存为 None，由后续按需逻辑处理。
    try:
        prefetch_bars(df_run["symbol"].tolist(), daily_period="1y", h4_period="90d")
    except Exception as exc:
        print(f"⚠️ 行情预取阶段出错（忽略，回退为按需下载）：{exc}", flush=True)

    xl = XunLongIndicator()

    all_rows = []
    scan_errors = []
    total = len(df_run)
    print(f"本次扫描标的数：{total}")
    for i, (_, r) in enumerate(df_run.iterrows(), start=1):
        sym = r["symbol"]
        name = r.get("name", "")
        pct = i / total * 100 if total else 100
        print(f"[{i}/{total} | {pct:.0f}%] 扫描 {sym} ({name})")
        try:
            df_sig = scan_one_symbol(sym, name, xl)
        except Exception as e:
            print(f"[{sym}] 扫描出错：{e}")
            traceback.print_exc()   # ← 多打印完整调用栈
            scan_errors.append((sym, str(e)))
            continue

        if not df_sig.empty:
            all_rows.append(df_sig)

    if scan_errors:
        print(
            f"⚠️ 本次有 {len(scan_errors)} 只标的扫描失败（数据缺失或异常），"
            f"不应与“无信号”混为一谈：",
            flush=True,
        )
        for esym, emsg in scan_errors[:20]:
            print(f"   - {esym}: {emsg}", flush=True)
        if len(scan_errors) > 20:
            print(f"   …以及另外 {len(scan_errors) - 20} 只", flush=True)

    forced_dates = _get_forced_rescan_signal_dates(run_dt)

    if not all_rows:
        print("本次没有任何股票触发信号。将生成空 Summary，并继续更新历史追踪sheet。")
        df_all = pd.DataFrame(columns=[
            "symbol", "name", "signal_date", "signal_type", "signal_side", "model", "close", "volume",
            "vol_ma20", "L2_trend", "L2_pump", "RSI", "rank120", "extra_info", "板块",
        ])
    else:
        df_all = pd.concat(all_rows, ignore_index=True)
        profile_map = df_run[["symbol", "group"]].drop_duplicates(subset=["symbol"], keep="first").copy()
        profile_map["板块"] = profile_map["group"].apply(_normalize_sector_with_code)
        profile_map = profile_map[["symbol", "板块"]]
        df_all = df_all.merge(profile_map, how="left", on="symbol")

        # 自动补回遗漏交易日的信号，避免漏跑后丢样本。
        # 手动重建时可用 STOCK_ONECLICK_RESCAN_FROM=YYYY-MM-DD 强制保留一段历史信号。
        catchup_dates = forced_dates or _get_catchup_signal_dates(HISTORY_DIR, run_dt, max_bdays=GANN_LOOKBACK_DAYS)
        df_all["signal_date"] = pd.to_datetime(df_all["signal_date"], errors="coerce").dt.date
        df_all = df_all[df_all["signal_date"].isin(catchup_dates)]
        df_all = (
            df_all.sort_values(["signal_date", "symbol", "signal_type"], ascending=[True, True, True])
            .drop_duplicates(subset=["symbol", "signal_type", "signal_date"], keep="first")
            .reset_index(drop=True)
        )
        if catchup_dates:
            label = "本次强制重扫信号日期" if forced_dates else "本次补回信号日期"
            print(f"{label}：", ", ".join([d.isoformat() for d in catchup_dates]))

    df_all["signal_date"] = pd.to_datetime(df_all["signal_date"], errors="coerce").dt.date
    df_all = df_all[df_all["signal_date"] >= LIFECYCLE_START_DATE].reset_index(drop=True)
    df_all = _drop_symbols(df_all, excluded_symbols)

    # 6) 加运行时间戳
    now = run_dt
    run_date = now.date()
    run_time = now.strftime("%H:%M:%S")
    df_all["run_date"] = run_date
    df_all["run_time"] = run_time
    df_all["buy_score"] = df_all.apply(score_buy_signal_row, axis=1) if not df_all.empty else np.nan
    df_all["sell_score"] = df_all.apply(score_sell_signal_row, axis=1) if not df_all.empty else np.nan

    col_order = [
        "run_date", "run_time",
        "symbol", "name", "板块",
        "signal_date", "signal_type", "signal_side", "model",
        "close", "volume", "vol_ma20",
        "L2_trend", "L2_pump", "RSI",
        "rank120", "H4_RSI", "H4_FJ", "H4_0_birth", "H4_1_birth",
        "Gann_1_date", "Gann_1_price", "buy_score", "sell_score", "extra_info",
    ]
    for c in col_order:
        if c not in df_all.columns:
            df_all[c] = np.nan
    formal_sell_mask = df_all["signal_type"].astype(str).eq("正式卖出")
    if formal_sell_mask.any():
        one_date_from_text = df_all.loc[formal_sell_mask, "extra_info"].astype(str).str.extract(r"1出日期=([0-9]{4}-[0-9]{2}-[0-9]{2})", expand=False)
        one_price_from_text = pd.to_numeric(
            df_all.loc[formal_sell_mask, "extra_info"].astype(str).str.extract(r"1出价格=([0-9]+(?:\.[0-9]+)?)", expand=False),
            errors="coerce",
        )
        missing_one_date = formal_sell_mask & pd.to_datetime(df_all["Gann_1_date"], errors="coerce").isna()
        missing_one_price = formal_sell_mask & pd.to_numeric(df_all["Gann_1_price"], errors="coerce").isna()
        df_all.loc[missing_one_date, "Gann_1_date"] = pd.to_datetime(
            one_date_from_text.reindex(df_all.index), errors="coerce"
        ).dt.date
        df_all.loc[missing_one_price, "Gann_1_price"] = one_price_from_text.reindex(df_all.index)
    if "板块" in df_all.columns:
        df_all["板块"] = df_all["板块"].apply(_normalize_sector_with_code)
    df_all = df_all[col_order]
    first_buy_count = int(df_all["signal_type"].astype(str).eq("第一买入点").sum())
    second_buy_count = int(df_all["signal_type"].astype(str).eq("二进宫买入点").sum())
    print(f"✅ 低位启动器：第一买入点 {first_buy_count} 条，二进宫买入点 {second_buy_count} 条", flush=True)

    # 追踪表：信号后 20 交易日表现（按 signal_date 建 sheet）
    print("阶段 4/4：生成信号后表现追踪sheet...", flush=True)
    history_source_dir = HISTORY_DIR
    if forced_dates:
        history_source_dir = HISTORY_DIR / ".rescan_current_only"
        history_source_dir.mkdir(parents=True, exist_ok=True)
        print("强制重扫模式：追踪表仅使用本次扫描结果，不合并旧 history。", flush=True)

    buy_anchors = _extract_anchor_signals(history_source_dir, df_all, signal_side="BUY")
    buy_anchors = _drop_symbols(buy_anchors, excluded_symbols)
    followup_sheets, completed_sheets = _build_followup_sheets(
        buy_anchors, now, max_days=TRACK_MAX_DAYS, sector_map=sector_map_all
    )
    sell_anchors = _extract_anchor_signals(history_source_dir, df_all, signal_side="SELL")
    sell_anchors = _drop_symbols(sell_anchors, excluded_symbols)
    sell_followup_sheets, sell_completed_sheets = _build_followup_sheets(
        sell_anchors, now, max_days=TRACK_MAX_DAYS, sector_map=sector_map_all, sheet_prefix="SELL_"
    )
    summary_buy_df = _build_active_section_summary(followup_sheets, df_run, now, signal_side="BUY", df_current_signals=df_all)
    summary_sell_df = _build_active_section_summary(sell_followup_sheets, df_run, now, signal_side="SELL", df_current_signals=df_all)
    buy_observation_df, buy_history_df, sell_observation_df, sell_history_df = _build_lifecycle_tables(
        history_source_dir, df_all, now, df_run, min_days=TRACK_MAX_DAYS
    )
    buy_observation_df = _drop_symbols(buy_observation_df, excluded_symbols)
    buy_history_df = _drop_symbols(buy_history_df, excluded_symbols)
    sell_observation_df = _drop_symbols(sell_observation_df, excluded_symbols)
    sell_history_df = _drop_symbols(sell_history_df, excluded_symbols)
    # tv_today 只导出“当天 D0 批次”的股票（与当日日期 sheet 保持一致）
    today_key = now.date().isoformat()
    if today_key in followup_sheets and not followup_sheets[today_key].empty:
        tv_symbols = (
            followup_sheets[today_key]["symbol"]
            .astype(str)
            .str.strip()
            .str.upper()
            .drop_duplicates()
        )
        tv_export_df = pd.DataFrame({"symbol": tv_symbols})
    else:
        tv_export_df = pd.DataFrame(columns=["symbol"])
    tv_export_df = tv_export_df.merge(
        df_run[["symbol", "exchange"]].drop_duplicates(subset=["symbol"], keep="first"),
        how="left",
        on="symbol",
    )
    combined_completed_dates = sorted(set(completed_sheets.keys()) | set(k.replace("SELL_", "") for k in sell_completed_sheets.keys()))
    market_context_cache = {}

    def market_context_for_sheet(dt_key: str):
        raw_key = str(dt_key).replace("SELL_", "")
        if raw_key in market_context_cache:
            return market_context_cache[raw_key]
        try:
            ctx_dt = pd.to_datetime(raw_key, errors="coerce")
            if pd.isna(ctx_dt):
                ctx_dt = pd.Timestamp(now)
            ctx = build_market_context(ctx_dt.to_pydatetime())
        except Exception as exc:
            ctx = None
            print(f"⚠️ 市场环境计算失败({raw_key})：{exc}", flush=True)
        market_context_cache[raw_key] = ctx
        return ctx

    today_market_context = market_context_for_sheet(today_key)
    if today_market_context:
        print(
            f"✅ 市场环境：{today_market_context.get('state')} | {today_market_context.get('daily_reason')}",
            flush=True,
        )

    archived_files = []
    for dt_key in combined_completed_dates:
        buy_completed_df = completed_sheets.get(dt_key, pd.DataFrame())
        sell_completed_df = sell_completed_sheets.get(f"SELL_{dt_key}", pd.DataFrame())
        date_key = dt_key.replace("-", "")
        out_file = COMPLETED_20D_DIR / f"{date_key}共计20天数据.xlsx"
        if out_file.exists():
            continue
        COMPLETED_20D_DIR.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            _write_combined_followup_sheet(
                writer,
                dt_key,
                buy_completed_df,
                sell_completed_df,
                market_context=market_context_for_sheet(dt_key),
            )
        archived_files.append(out_file)

    if "板块" in summary_buy_df.columns:
        summary_buy_df["板块"] = summary_buy_df["板块"].apply(_normalize_sector_with_code)
    if "板块" in summary_sell_df.columns:
        summary_sell_df["板块"] = summary_sell_df["板块"].apply(_normalize_sector_with_code)

    recent_result_map = {} if forced_dates else _recent_history_result_map(HISTORY_DIR, now, TRACK_MAX_DAYS)

    # 历史归档：每次都保存一份到 history，主目录保留一份 latest 便于打开
    out_name = f"scan_result_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    history_path = HISTORY_DIR / out_name
    latest_path = BASE_DIR / "scan_result_latest.xlsx"
    with pd.ExcelWriter(history_path, engine="openpyxl") as writer:
        _write_summary_sheet(writer, summary_buy_df, summary_sell_df)
        _write_raw_signals_sheet(writer, df_all)
        _write_simple_table_sheet(writer, "买入观察列表", buy_observation_df)
        _write_simple_table_sheet(writer, "买入历史记录", buy_history_df)
        _write_simple_table_sheet(writer, "卖出观察列表", sell_observation_df)
        _write_simple_table_sheet(writer, "卖出历史记录", sell_history_df)
        combined_dates = sorted(
            set(followup_sheets.keys())
            | set(k.replace("SELL_", "") for k in sell_followup_sheets.keys())
            | set(recent_result_map.keys())
            | {today_key},
            reverse=True,
        )
        for dt_key in combined_dates:
            buy_sdf = followup_sheets.get(dt_key, pd.DataFrame())
            sell_sdf = sell_followup_sheets.get(f"SELL_{dt_key}", pd.DataFrame())
            if not buy_sdf.empty or not sell_sdf.empty:
                _write_combined_followup_sheet(
                    writer,
                    dt_key,
                    buy_sdf,
                    sell_sdf,
                    market_context=market_context_for_sheet(dt_key),
                )
                continue

            if dt_key == today_key:
                buy_snapshot_df = df_all[df_all["signal_side"].astype(str).str.upper() == "BUY"].copy()
                sell_snapshot_df = df_all[df_all["signal_side"].astype(str).str.upper() == "SELL"].copy()
            else:
                hist_path = recent_result_map.get(dt_key)
                buy_snapshot_df = _read_signal_rows_from_result(hist_path, "BUY") if hist_path else pd.DataFrame()
                sell_snapshot_df = _read_signal_rows_from_result(hist_path, "SELL") if hist_path else pd.DataFrame()
            buy_snapshot_df = _drop_symbols(buy_snapshot_df, excluded_symbols)
            sell_snapshot_df = _drop_symbols(sell_snapshot_df, excluded_symbols)

            _write_combined_snapshot_sheet(
                writer,
                dt_key,
                buy_snapshot_df,
                sell_snapshot_df,
                market_context=market_context_for_sheet(dt_key),
            )

    shutil.copy2(history_path, latest_path)
    dashboard_path, dashboard_hist_path = export_signal_dashboard(df_all, now, EXPORT_DIR, HISTORY_DIR)
    print(f"✅ 历史结果已保存：{history_path}")
    print(f"✅ 最新结果已更新：{latest_path}")
    print(f"✅ 信号Dashboard：{dashboard_path}")
    print(f"✅ 信号Dashboard历史：{dashboard_hist_path}")
    print(f"✅ 买入追踪sheet数量：{len(followup_sheets)}（每批次最多追踪 {TRACK_MAX_DAYS} 个交易日）")
    print(f"✅ 卖出追踪批次数：{len(sell_followup_sheets)}（每批次最多追踪 {TRACK_MAX_DAYS} 个交易日）")
    print(f"✅ 买入观察列表：{len(buy_observation_df)} 条")
    print(f"✅ 买入历史记录：{len(buy_history_df)} 条")
    print(f"✅ 卖出观察列表：{len(sell_observation_df)} 条")
    print(f"✅ 卖出历史记录：{len(sell_history_df)} 条")
    if archived_files:
        print(f"✅ 已归档满{TRACK_MAX_DAYS}日批次：{len(archived_files)} 个")
        for p in archived_files:
            print(f"   - {p}")

    # 6) TradingView 每日清单（按日期命名）+ 历史清单
    tv_today_path, tv_hist_path = export_tradingview_lists(tv_export_df, now, EXPORT_DIR, HISTORY_DIR)
    tv_buy_path, tv_buy_notes_path, tv_buy_latest, tv_buy_notes_latest = export_tv_buy_signal_notes(
        followup_sheets.get(today_key, pd.DataFrame()),
        now,
        df_run,
        EXPORT_DIR,
        HISTORY_DIR,
    )
    print(f"✅ TV今日清单：{tv_today_path}")
    print(f"✅ TV历史清单：{tv_hist_path}")
    print(f"✅ TV买入纯导入：{tv_buy_path}")
    print(f"✅ TV买入备注版：{tv_buy_notes_path}")

    # 自动打开结果（Mac）。设置 STOCK_ONECLICK_NO_OPEN=1 可跳过（无人值守 / CI）。
    if os.environ.get("STOCK_ONECLICK_NO_OPEN", "").strip() == "1":
        print("STOCK_ONECLICK_NO_OPEN=1：跳过自动打开结果。")
    else:
        try:
            subprocess.run(["open", str(latest_path)])
        except Exception as e:
            print("自动打开结果失败，但文件已保存：", e)


if __name__ == "__main__":
    main()
